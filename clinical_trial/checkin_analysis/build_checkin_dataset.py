from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict, deque
from copy import copy
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_USER_COLUMNS = ["user_nickname", "participant_nickname", "用户昵称"]
DEFAULT_SENDER_COLUMNS = ["sender_name", "发送者名称"]
DEFAULT_SENDER_NICKNAME_COLUMNS = ["sender_group_nickname", "发送者群昵称(未修改则为空)"]
DEFAULT_CONTENT_COLUMNS = ["message_content", "消息内容"]
BLANK_STREAK_LIMIT = 200

INVISIBLE_SPACES = "\u2005\u2002\u2003\u3000\u00a0\u2009\u202f\ufeff\u200b"
SPACE_TRANSLATION = str.maketrans({char: " " for char in INVISIBLE_SPACES})
QUOTE_TRANSLATION = str.maketrans({"“": '"', "”": '"', "‘": "'", "’": "'"})
SCORE_SUFFIX_RE = re.compile(r"[\s+\-_.～~一—]*\d+(?:[.．]\d+)?(?:[\s+\-_.～~一—]*\d+(?:[.．]\d+)?)*$")


def log(message: str) -> None:
    print(f"[checkin-build] {message}", flush=True)


def ensure_parent_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def normalize_text(text: object) -> str:
    text = str(text).translate(SPACE_TRANSLATION).translate(QUOTE_TRANSLATION)
    return re.sub(r"[ \t\r\f\v]+", " ", text).strip()


def clean_name(name: object | None) -> str:
    if name is None:
        return ""
    return normalize_text(name).strip(" @\"'`:,;:：")


def alias_variants(name: object | None) -> set[str]:
    base = clean_name(name)
    if not base:
        return set()

    variants = {base, base.lower(), base.replace(" ", ""), base.replace(" ", "").lower()}
    score_stripped = SCORE_SUFFIX_RE.sub("", base).strip(" +-_.～~一—")
    if score_stripped:
        variants.update(
            {
                score_stripped,
                score_stripped.lower(),
                score_stripped.replace(" ", ""),
                score_stripped.replace(" ", "").lower(),
            }
        )

    alnum_only = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "", base)
    if alnum_only:
        variants.update({alnum_only, alnum_only.lower()})

    if score_stripped:
        alnum_score = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "", score_stripped)
        if alnum_score:
            variants.update({alnum_score, alnum_score.lower()})

    return {variant for variant in variants if variant}


def pick_header_name(
    header: Sequence[object],
    candidates: Sequence[str],
    *,
    required: bool,
    label: str,
    path: Path,
) -> str | None:
    exact = {str(value).strip(): str(value).strip() for value in header if value is not None}
    lower = {key.lower(): key for key in exact}
    for candidate in candidates:
        candidate_clean = str(candidate).strip()
        if candidate_clean in exact:
            return exact[candidate_clean]
        if candidate_clean.lower() in lower:
            return lower[candidate_clean.lower()]
    if required:
        raise KeyError(f"Could not find {label} column in {path.name}. Candidates={list(candidates)}")
    return None


def iter_effective_rows(worksheet, start_row: int = 2, key_col: int | None = None):
    blank_streak = 0
    for row in worksheet.iter_rows(min_row=start_row, values_only=True):
        key_has_value = False
        if key_col is not None and key_col < len(row):
            key_has_value = bool(clean_name(row[key_col]))
        row_has_value = any(value not in (None, "") for value in row)

        is_effective = key_has_value if key_col is not None else row_has_value
        if not is_effective:
            blank_streak += 1
            if blank_streak >= BLANK_STREAK_LIMIT:
                break
            continue

        blank_streak = 0
        yield row


def detect_last_data_row(worksheet, key_col_1based: int) -> int:
    last_row = 1
    blank_streak = 0
    key_idx = key_col_1based - 1
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        has_key = key_idx < len(row) and bool(clean_name(row[key_idx]))
        has_any_value = any(value not in (None, "") for value in row)
        if has_key:
            last_row = row_index
            blank_streak = 0
            continue

        if not has_any_value:
            blank_streak += 1
            if blank_streak >= BLANK_STREAK_LIMIT:
                break
        else:
            blank_streak = 0
    return last_row


def clipped_auto_filter_ref(ref: str | None, last_row: int) -> str | None:
    if not ref:
        return ref
    match = re.fullmatch(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", ref)
    if not match:
        return ref
    start_col, start_row, end_col, _ = match.groups()
    end_row = max(last_row, int(start_row))
    return f"{start_col}{start_row}:{end_col}{end_row}"


def build_main_variant_map(users: list[str]) -> dict[str, str]:
    variant_map: dict[str, set[str]] = defaultdict(set)
    for user in users:
        for variant in alias_variants(user):
            variant_map[variant].add(user)
    return {variant: next(iter(targets)) for variant, targets in variant_map.items() if len(targets) == 1}


def build_alias_graph(chat_path: Path, sender_col: str, nickname_col: str | None) -> dict[str, set[str]]:
    workbook = load_workbook(chat_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header = [str(value) if value is not None else "" for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    sender_idx = header.index(sender_col)
    nickname_idx = header.index(nickname_col) if nickname_col else None

    graph: dict[str, set[str]] = defaultdict(set)
    for row in iter_effective_rows(worksheet, start_row=2):
        aliases: list[str] = []
        values = [row[sender_idx]]
        if nickname_idx is not None:
            values.append(row[nickname_idx])
        for value in values:
            cleaned = clean_name(value)
            if cleaned:
                aliases.append(cleaned)
        aliases = list(dict.fromkeys(aliases))
        for alias in aliases:
            graph.setdefault(alias, set()).add(alias)
        for alias in aliases:
            graph[alias].update(aliases)

    workbook.close()
    return graph


def connected_aliases(alias: str, graph: dict[str, set[str]]) -> set[str]:
    alias = clean_name(alias)
    if not alias:
        return set()
    if alias not in graph:
        return {alias}

    seen = set()
    queue = deque([alias])
    while queue:
        current = queue.popleft()
        if current in seen:
            continue
        seen.add(current)
        for neighbor in graph.get(current, set()):
            if neighbor not in seen:
                queue.append(neighbor)
    return seen


def resolve_alias(
    alias: str,
    *,
    manual_aliases: dict[str, str],
    main_users: set[str],
    main_variant_map: dict[str, str],
    graph: dict[str, set[str]],
) -> str | None:
    alias = clean_name(alias)
    if not alias:
        return None

    manual_target = manual_aliases.get(alias)
    if manual_target in main_users:
        return manual_target

    candidates = set()
    for related in connected_aliases(alias, graph):
        for variant in alias_variants(related):
            target = main_variant_map.get(variant)
            if target:
                candidates.add(target)
        manual_related = manual_aliases.get(related)
        if manual_related in main_users:
            candidates.add(manual_related)

    if len(candidates) == 1:
        return next(iter(candidates))
    return None


def extract_target_candidates(message: str) -> list[str]:
    text = normalize_text(message)
    lines = [line.strip() for line in str(message).translate(SPACE_TRANSLATION).translate(QUOTE_TRANSLATION).splitlines() if line.strip()]
    candidates: list[str] = []

    for match in re.finditer(r"@([^\s#]+)", text):
        candidates.append(match.group(1))

    if lines:
        match = re.match(r'^"?([^:："]+)[:：]', lines[0])
        if match:
            candidates.append(match.group(1))

    deduplicated: list[str] = []
    seen = set()
    for candidate in candidates:
        cleaned = clean_name(candidate)
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            deduplicated.append(cleaned)
    return deduplicated


def load_main_users(main_path: Path, user_column: str) -> tuple[list[str], int]:
    workbook = load_workbook(main_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header = [str(value) if value is not None else "" for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    user_col = header.index(user_column)

    users: list[str] = []
    for row in iter_effective_rows(worksheet, start_row=2, key_col=user_col):
        value = clean_name(row[user_col])
        if value:
            users.append(value)

    workbook.close()
    return users, user_col + 1


def build_counts(
    *,
    users: list[str],
    chat_path: Path,
    sender_column: str,
    sender_nickname_column: str | None,
    content_column: str,
    keyword: str,
    manual_aliases: dict[str, str],
) -> tuple[Counter, dict]:
    main_users = {user for user in users if user}
    main_variant_map = build_main_variant_map([user for user in users if user])

    log(f"Building alias graph from {chat_path.name} ...")
    graph = build_alias_graph(chat_path, sender_column, sender_nickname_column)

    workbook = load_workbook(chat_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header = [str(value) if value is not None else "" for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    content_idx = header.index(content_column)

    counts: Counter = Counter()
    alias_usage: dict[str, set[str]] = defaultdict(set)
    unresolved_examples = []
    ambiguous_examples = []
    keyword_message_count = 0
    resolved_message_count = 0

    log(f"Scanning keyword messages containing {keyword!r} ...")
    for row in iter_effective_rows(worksheet, start_row=2):
        content = row[content_idx]
        if content is None:
            continue
        message = str(content)
        if keyword not in message:
            continue

        keyword_message_count += 1
        candidates = extract_target_candidates(message)
        resolved_users = []
        for candidate in candidates:
            resolved = resolve_alias(
                candidate,
                manual_aliases=manual_aliases,
                main_users=main_users,
                main_variant_map=main_variant_map,
                graph=graph,
            )
            if resolved:
                alias_usage[candidate].add(resolved)
                resolved_users.append(resolved)

        unique_users = list(dict.fromkeys(resolved_users))
        if len(unique_users) == 1:
            counts[unique_users[0]] += 1
            resolved_message_count += 1
        elif len(unique_users) > 1:
            ambiguous_examples.append(
                {
                    "candidates": candidates,
                    "resolved_users": unique_users,
                    "message_preview": message[:240],
                }
            )
        else:
            unresolved_examples.append(
                {
                    "candidates": candidates,
                    "message_preview": message[:240],
                }
            )

    workbook.close()
    log(
        f"Matched {resolved_message_count}/{keyword_message_count} keyword messages; "
        f"unresolved={len(unresolved_examples)}, ambiguous={len(ambiguous_examples)}."
    )

    report = {
        "keyword": keyword,
        "main_user_count": len(main_users),
        "keyword_message_count": keyword_message_count,
        "resolved_message_count": resolved_message_count,
        "unresolved_message_count": len(unresolved_examples),
        "ambiguous_message_count": len(ambiguous_examples),
        "users_with_nonzero_count": [{"user": user, "count": count} for user, count in counts.most_common()],
        "alias_resolution_examples": {alias: sorted(targets) for alias, targets in sorted(alias_usage.items())},
        "unresolved_examples": unresolved_examples[:20],
        "ambiguous_examples": ambiguous_examples[:20],
    }
    return counts, report


def clone_adjacent_style(worksheet, source_col: int, target_col: int, last_row: int) -> None:
    source_width = worksheet.column_dimensions[get_column_letter(source_col)].width
    if source_width is not None:
        worksheet.column_dimensions[get_column_letter(target_col)].width = source_width

    for row_index in range(1, last_row + 1):
        source_cell = worksheet.cell(row=row_index, column=source_col)
        target_cell = worksheet.cell(row=row_index, column=target_col)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)


def copy_source_sheet(main_file: Path, user_column: str) -> tuple[Workbook, object, int]:
    source_workbook = load_workbook(main_file, data_only=False)
    source_worksheet = source_workbook[source_workbook.sheetnames[0]]
    header = [source_worksheet.cell(row=1, column=column).value for column in range(1, source_worksheet.max_column + 1)]
    user_col = header.index(user_column) + 1
    last_row = detect_last_data_row(source_worksheet, user_col)

    target_workbook = Workbook()
    target_worksheet = target_workbook.active
    target_worksheet.title = source_worksheet.title

    for row in source_worksheet.iter_rows(min_row=1, max_row=last_row):
        for source_cell in row:
            target_cell = target_worksheet.cell(row=source_cell.row, column=source_cell.column, value=source_cell.value)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format

    for column_letter, dimension in source_worksheet.column_dimensions.items():
        target_dimension = target_worksheet.column_dimensions[column_letter]
        target_dimension.width = dimension.width
        target_dimension.hidden = dimension.hidden
        target_dimension.bestFit = dimension.bestFit

    for row_index, dimension in source_worksheet.row_dimensions.items():
        if row_index > last_row:
            continue
        target_dimension = target_worksheet.row_dimensions[row_index]
        target_dimension.height = dimension.height
        target_dimension.hidden = dimension.hidden

    for merged_range in source_worksheet.merged_cells.ranges:
        target_worksheet.merge_cells(str(merged_range))

    target_worksheet.freeze_panes = source_worksheet.freeze_panes
    target_worksheet.auto_filter.ref = clipped_auto_filter_ref(source_worksheet.auto_filter.ref, last_row)

    source_workbook.close()
    return target_workbook, target_worksheet, last_row


def write_output(
    *,
    main_file: Path,
    user_column: str,
    new_column_title: str,
    output_file: Path,
    counts: Counter,
) -> None:
    workbook, worksheet, last_row = copy_source_sheet(main_file, user_column)
    header = [worksheet.cell(row=1, column=column).value for column in range(1, worksheet.max_column + 1)]
    user_col = header.index(user_column) + 1
    new_col = worksheet.max_column + 1

    clone_adjacent_style(worksheet, new_col - 1, new_col, last_row)
    worksheet.cell(row=1, column=new_col).value = new_column_title
    worksheet.cell(row=1, column=new_col).number_format = "General"
    worksheet.column_dimensions[get_column_letter(new_col)].width = max(len(new_column_title) + 2, 18)

    log(f"Writing count column for {last_row - 1} data rows ...")
    for row_index in range(2, last_row + 1):
        user = clean_name(worksheet.cell(row=row_index, column=user_col).value)
        worksheet.cell(row=row_index, column=new_col).value = counts.get(user, 0)
        worksheet.cell(row=row_index, column=new_col).number_format = "0"

    ensure_parent_dir(output_file)
    log(f"Saving workbook to {output_file} ...")
    workbook.save(output_file)
    workbook.close()


def derive_report_path(output_file: Path, report_file: Path | None) -> Path:
    if report_file is not None:
        return report_file
    return output_file.with_name(f"{output_file.stem}_report.json")


def load_manual_aliases(path: Path | None) -> dict[str, str]:
    if path is None:
        return {}
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError(f"Alias mapping file must contain a JSON object: {path}")
    cleaned: dict[str, str] = {}
    for alias, target in raw.items():
        alias_clean = clean_name(alias)
        target_clean = clean_name(target)
        if alias_clean and target_clean:
            cleaned[alias_clean] = target_clean
    return cleaned


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Add a chat-derived actual feedback-count column to a clinical-trial workbook."
    )
    parser.add_argument("--main-file", type=Path, required=True, help="Participant workbook to augment.")
    parser.add_argument("--chat-file", type=Path, required=True, help="Chat-export workbook used to count tagged messages.")
    parser.add_argument("--output-file", type=Path, required=True, help="Destination workbook with the appended count column.")
    parser.add_argument("--report-file", type=Path, default=None, help="Optional JSON report path. Defaults to <output-file stem>_report.json.")
    parser.add_argument("--keyword", required=True, help="Keyword or hashtag that identifies the messages to count.")
    parser.add_argument("--new-column-title", required=True, help="Header for the appended count column.")
    parser.add_argument("--user-column", default=None, help="Canonical participant-name column in the main workbook.")
    parser.add_argument("--sender-column", default=None, help="Sender-name column in the chat workbook.")
    parser.add_argument("--sender-nickname-column", default=None, help="Optional sender group-nickname column in the chat workbook.")
    parser.add_argument("--content-column", default=None, help="Message-content column in the chat workbook.")
    parser.add_argument("--alias-json", type=Path, default=None, help="Optional JSON mapping of alias -> canonical participant name.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    report_file = derive_report_path(args.output_file, args.report_file)
    manual_aliases = load_manual_aliases(args.alias_json)

    main_workbook = load_workbook(args.main_file, read_only=True, data_only=True)
    main_sheet = main_workbook[main_workbook.sheetnames[0]]
    main_header = [str(value) if value is not None else "" for value in next(main_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    user_column = pick_header_name(
        main_header,
        [args.user_column] if args.user_column else DEFAULT_USER_COLUMNS,
        required=True,
        label="participant-name",
        path=args.main_file,
    )
    main_workbook.close()

    chat_workbook = load_workbook(args.chat_file, read_only=True, data_only=True)
    chat_sheet = chat_workbook[chat_workbook.sheetnames[0]]
    chat_header = [str(value) if value is not None else "" for value in next(chat_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    sender_column = pick_header_name(
        chat_header,
        [args.sender_column] if args.sender_column else DEFAULT_SENDER_COLUMNS,
        required=True,
        label="sender-name",
        path=args.chat_file,
    )
    sender_nickname_column = pick_header_name(
        chat_header,
        [args.sender_nickname_column] if args.sender_nickname_column else DEFAULT_SENDER_NICKNAME_COLUMNS,
        required=False,
        label="sender-group-nickname",
        path=args.chat_file,
    )
    content_column = pick_header_name(
        chat_header,
        [args.content_column] if args.content_column else DEFAULT_CONTENT_COLUMNS,
        required=True,
        label="message-content",
        path=args.chat_file,
    )
    chat_workbook.close()

    log(f"Loading users from {args.main_file.name} ...")
    users, _ = load_main_users(args.main_file, user_column)
    log(f"Loaded {len(users)} participants.")

    counts, report = build_counts(
        users=users,
        chat_path=args.chat_file,
        sender_column=sender_column,
        sender_nickname_column=sender_nickname_column,
        content_column=content_column,
        keyword=args.keyword,
        manual_aliases=manual_aliases,
    )
    write_output(
        main_file=args.main_file,
        user_column=user_column,
        new_column_title=args.new_column_title,
        output_file=args.output_file,
        counts=counts,
    )

    report.update(
        {
            "main_file": str(args.main_file),
            "chat_file": str(args.chat_file),
            "output_file": str(args.output_file),
            "report_file": str(report_file),
            "user_column": user_column,
            "sender_column": sender_column,
            "sender_nickname_column": sender_nickname_column,
            "content_column": content_column,
            "new_column_title": args.new_column_title,
            "manual_alias_count": len(manual_aliases),
        }
    )

    ensure_parent_dir(report_file)
    report_file.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"Wrote report to {report_file}.")


if __name__ == "__main__":
    main()
