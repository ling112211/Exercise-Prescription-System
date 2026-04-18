from __future__ import annotations

import argparse
import contextlib
import io
import json
import math
import re
from collections import defaultdict, deque
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from statistics import NormalDist
from typing import Any

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HUMAN_GROUP = 0.0
EPS_GROUP = 1.0
DEFAULT_N_BOOT = 5000
DEFAULT_SEED = 42
STANDARD_NORMAL = NormalDist()

USER_COLUMNS = ["user_nickname", "participant_nickname", "用户昵称"]
AGE_COLUMNS = ["age", "年龄"]
SEX_COLUMNS = ["sex", "gender", "性别"]
BMI_COLUMNS = ["BMI", "bmi", "体重指数"]
HUMAN_COUNT_COLUMNS = ["daily_activity_checkin_count", "exercise_feedback_count", "包含#日常活动打卡消息总数"]
EPS_COUNT_COLUMNS = ["exercise_feedback_count", "包含#运动点评消息总数"]

WEIGHT_BASELINE_COLUMNS = ["baseline_weight_kg", "baseline_weight", "weight_baseline", "入营体重", "入营体重kg", "入营体重（档案）", "入营体重（档案） "]
WEIGHT_ENDLINE_COLUMNS = ["endline_weight_kg", "endline_weight", "weight_endline", "出营体重", "出营体重kg"]
WEIGHT_LOSS_KG_COLUMNS = ["weight_loss_kg", "减重数"]
WEIGHT_LOSS_PCT_COLUMNS = ["weight_loss_pct", "减重比率"]

GLYCEMIC_BASELINE_COLUMNS = ["baseline_fpg_mmol", "baseline_fpg", "FPG0", "入营空腹", "Baseline fasting glucose", "Fasting glucose"]
GLYCEMIC_ENDLINE_COLUMNS = ["endpoint_fpg_mmol", "endline_fpg_mmol", "endpoint_fpg", "结营空腹"]

CHAT_TIMESTAMP_COLUMNS = ["timestamp", "发送时间"]
CHAT_SENDER_COLUMNS = ["sender_name", "发送者名称"]
CHAT_NICKNAME_COLUMNS = ["sender_group_nickname", "发送者群昵称(未修改则为空)"]
CHAT_CONTENT_COLUMNS = ["message_content", "消息内容"]

LINE_FILL = PatternFill("solid", start_color="1D4ED8")
SECTION_FILL = PatternFill("solid", start_color="EFF6FF")
WARN_FILL = PatternFill("solid", start_color="FEF3C7")
THIN = Side(style="thin", color="CBD5E1")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, color="1D4ED8")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

INVISIBLE_SPACES = "\u2005\u2002\u2003\u3000\u00a0\u2009\u202f\ufeff\u200b"
SPACE_TRANSLATION = str.maketrans({char: " " for char in INVISIBLE_SPACES})
QUOTE_TRANSLATION = str.maketrans({"“": '"', "”": '"', "‘": "'", "’": "'", "\u201c": '"', "\u201d": '"', "\u2018": "'", "\u2019": "'"})
SCORE_SUFFIX_RE = re.compile(r"[\s+\-_.～~一—]*\d+(?:[.．]\d+)?(?:[\s+\-_.～~一—]*\d+(?:[.．]\d+)?)*$")

# The keyword dictionaries below are intentionally tied to the observed feedback
# segments. Human-arm records can bundle exercise feedback with weight and meal
# comments, so quality coding first isolates the requested feedback hashtag
# segment before scoring these features.
EXERCISE_PARAM_RE = re.compile(
    r"心率|次/分钟|分钟|小时|公里|千米|米|步|步数|卡路里|千卡|热量|消耗|"
    r"MET|强度|综合评分|评分|时长|配速|速度|距离|运动时间|有氧|无氧|"
    r"耐力|力量训练|抗阻|拉伸|热身|训练量|运动量|跑步|慢跑|快走|"
    r"步行|散步|游泳|骑行|瑜伽|HIIT|"
    r"steps?|step count|walk(?:ing)?|run(?:ning)?|jog(?:ging)?|cycling|bike|yoga|"
    r"stretch(?:ing)?|calories?|pace|duration|minutes?|heart rate|intensity",
    re.IGNORECASE,
)
PERSONALISE_RE = re.compile(
    r"根据.{0,12}(运动记录|打卡|记录|数据)|"
    r"结合.{0,12}(主观感受|心率|情况|记录|数据)|"
    r"您在本次|你在本次|本次运动|您的运动记录|你的运动记录|"
    r"从数据来看|针对你|针对您|你的情况|您的情况|你目前|您目前|"
    r"上次|上一次|最近你|最近您|你上次|您上次|"
    r"based on your|for you|your current|your recent|last time|last week|this week|today you",
    re.IGNORECASE,
)
ADVICE_RE = re.compile(
    r"建议|可以尝试|推荐|注意|下次|下周|调整|增加|减少|降低|提高|"
    r"适当|安排|尝试|补充水分|休息|拉伸|热身|控制|避免|循序渐进|"
    r"suggest|recommend|try|adjust|increase|decrease|keep|continue|consider|aim for",
    re.IGNORECASE,
)
ENCOURAGEMENT_RE = re.compile(
    r"加油|太棒|非常棒|很棒|不错|赞|厉害|棒棒|做得好|辛苦了|"
    r"非常好|太好了|进步|优秀|继续保持|值得认可|佩服|出色|满分|坚持|"
    r"great|nice job|well done|excellent|good work|keep it up|awesome|amazing",
    re.IGNORECASE,
)
GOAL_PLAN_RE = re.compile(
    r"目标|计划|每周|每次|下次|安排|逐渐|持续|坚持|习惯|实现|达到|"
    r"减脂|减肥|体重管理|心肺功能|耐力|力量训练|运动量|训练|冲刺|挑战|"
    r"goal|plan|routine|weekly|next session|next week|habit|target",
    re.IGNORECASE,
)
DATA_REFERENCE_RE = re.compile(
    r"根据您的运动记录|根据你的运动记录|从数据来看|结合您的主观感受|"
    r"结合你的主观感受|主观感受|平均心率|心率百分比|本次消耗|"
    r"综合评分|评分为|相当于|进行了|运动记录|数据来看|记录显示|运动表现|"
    r"record shows|your data|based on.*record|average heart rate|calories burned",
    re.IGNORECASE,
)
POSITIVE_RE = ENCOURAGEMENT_RE

QUALITY_BINARY_FEATURES: list[tuple[str, str]] = [
    ("has_personalisation", "Personalisation"),
    ("has_exercise_params", "Exercise parameters"),
    ("has_advice", "Specific advice"),
    ("has_encouragement", "Encouragement"),
    ("has_goal_plan", "Goal/plan language"),
    ("has_data_reference", "Data references"),
]


@dataclass(frozen=True)
class OutcomeSpec:
    key: str
    title: str
    description: str
    note: str
    covariate_key: str


@dataclass(frozen=True)
class CohortConfig:
    key: str
    label: str
    title: str
    human_keyword: str
    eps_keyword: str
    primary_plot_outcome: str
    plot_ylabel: str
    outcome_specs: tuple[OutcomeSpec, ...]


COHORT_CONFIGS: dict[str, CohortConfig] = {
    "weight_loss": CohortConfig(
        key="weight_loss",
        label="weight-loss cohort",
        title="Frequency-control and feedback-content audit - weight-loss cohort",
        human_keyword="#日常活动打卡",
        eps_keyword="#运动点评",
        primary_plot_outcome="weight_loss_kg",
        plot_ylabel="Weight loss (kg)",
        outcome_specs=(
            OutcomeSpec(
                key="endline_weight",
                title="Primary ANCOVA Endpoint",
                description="Outcome = day-21 body weight (kg); lower is better.",
                note="Negative arm effects favour EPS because they imply lower endline body weight after baseline adjustment.",
                covariate_key="baseline_weight",
            ),
            OutcomeSpec(
                key="weight_loss_kg",
                title="Weight-Loss (kg) Sensitivity",
                description="Outcome = baseline minus day-21 weight (kg); higher is better.",
                note="Positive arm effects favour EPS because they imply larger absolute weight loss.",
                covariate_key="baseline_weight",
            ),
            OutcomeSpec(
                key="weight_loss_pct",
                title="Percent Weight-Loss Sensitivity",
                description="Outcome = percent weight loss from baseline; higher is better.",
                note="Positive arm effects favour EPS because they imply larger percent weight loss.",
                covariate_key="baseline_weight",
            ),
        ),
    ),
    "glycemic": CohortConfig(
        key="glycemic",
        label="glycemic-control cohort",
        title="Frequency-control and feedback-content audit - glycemic-control cohort",
        human_keyword="#日常活动打卡",
        eps_keyword="#运动点评",
        primary_plot_outcome="fpg_reduction",
        plot_ylabel="FPG reduction (mmol/L)",
        outcome_specs=(
            OutcomeSpec(
                key="endline_fpg",
                title="Primary ANCOVA Endpoint",
                description="Outcome = day-21 fasting plasma glucose (mmol/L); lower is better.",
                note="Negative arm effects favour EPS because they imply lower endline fasting glucose after baseline adjustment.",
                covariate_key="baseline_fpg",
            ),
            OutcomeSpec(
                key="fpg_reduction",
                title="FPG Reduction Sensitivity",
                description="Outcome = baseline minus day-21 fasting plasma glucose (mmol/L); higher is better.",
                note="Positive arm effects favour EPS because they imply larger fasting-glucose reduction.",
                covariate_key="baseline_fpg",
            ),
        ),
    ),
}


def log(message: str) -> None:
    print(f"[checkin-frequency-control] {message}", flush=True)


def ensure_parent_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def normalize_text(text: object) -> str:
    text = str(text).translate(SPACE_TRANSLATION).translate(QUOTE_TRANSLATION)
    return re.sub(r"[ \t\r\f\v]+", " ", text).strip()


def clean_name(name: object | None) -> str:
    if name is None:
        return ""
    return normalize_text(name).strip(" @\"'`:,;:：")


def pick_header_name(
    header: list[str],
    candidates: list[str],
    *,
    required: bool,
    label: str,
    path: Path,
) -> str | None:
    exact = {str(value).strip(): str(value).strip() for value in header if value is not None}
    lower = {key.lower(): key for key in exact}
    for candidate in candidates:
        candidate_clean = str(candidate).strip()
        if candidate_clean in exact:
            return exact[candidate_clean]
        if candidate_clean.lower() in lower:
            return lower[candidate_clean.lower()]
    if required:
        raise KeyError(f"Could not find {label} column in {path.name}. Candidates={candidates}")
    return None


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("%", "")
    try:
        return float(text)
    except ValueError:
        return None


def to_clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def sex_to_female_indicator(value: Any) -> float | None:
    text = to_clean_text(value).strip().lower()
    if not text:
        return None
    female_tokens = {"f", "female", "woman", "女", "0", "2"}
    male_tokens = {"m", "male", "man", "男", "1"}
    if text in female_tokens:
        return 1.0
    if text in male_tokens:
        return 0.0
    return 0.0


def format_number(value: Any, digits: int = 4) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "NA"
    if isinstance(value, int):
        return str(value)
    return f"{value:.{digits}f}"


def format_pct(value: float | None) -> str:
    return "NA" if value is None else f"{value:.1%}"


def pstar(p_value: float | None) -> str:
    if p_value is None or math.isnan(p_value):
        return "NA"
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    return "ns"


def normal_two_sided_p(z_value: float | None) -> float | None:
    if z_value is None or math.isnan(z_value):
        return None
    return 2.0 * (1.0 - STANDARD_NORMAL.cdf(abs(z_value)))


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def alias_variants(name: object | None) -> set[str]:
    base = clean_name(name)
    if not base:
        return set()
    variants = {base, base.lower(), base.replace(" ", ""), base.replace(" ", "").lower()}
    score_stripped = SCORE_SUFFIX_RE.sub("", base).strip(" +-_.～~一—")
    if score_stripped:
        variants.update(
            {
                score_stripped,
                score_stripped.lower(),
                score_stripped.replace(" ", ""),
                score_stripped.replace(" ", "").lower(),
            }
        )
    alnum_only = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "", base)
    if alnum_only:
        variants.update({alnum_only, alnum_only.lower()})
    if score_stripped:
        alnum_score = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "", score_stripped)
        if alnum_score:
            variants.update({alnum_score, alnum_score.lower()})
    return {variant for variant in variants if variant}


def build_variant_map(users: list[str]) -> dict[str, str]:
    variant_map: dict[str, set[str]] = defaultdict(set)
    for user in users:
        for variant in alias_variants(user):
            variant_map[variant].add(user)
    return {variant: next(iter(targets)) for variant, targets in variant_map.items() if len(targets) == 1}


def build_alias_graph(chat_path: Path) -> dict[str, set[str]]:
    workbook = load_workbook(chat_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header = [str(value) if value is not None else "" for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    sender_column = pick_header_name(header, CHAT_SENDER_COLUMNS, required=False, label="sender", path=chat_path)
    nickname_column = pick_header_name(header, CHAT_NICKNAME_COLUMNS, required=False, label="sender-nickname", path=chat_path)
    sender_idx = header.index(sender_column) if sender_column else None
    nickname_idx = header.index(nickname_column) if nickname_column else None

    graph: dict[str, set[str]] = defaultdict(set)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in row):
            continue
        aliases: list[str] = []
        for idx in (sender_idx, nickname_idx):
            if idx is None:
                continue
            cleaned = clean_name(row[idx])
            if cleaned:
                aliases.append(cleaned)
        aliases = list(dict.fromkeys(aliases))
        for alias in aliases:
            graph.setdefault(alias, set()).add(alias)
        for alias in aliases:
            graph[alias].update(aliases)
    workbook.close()
    return graph


def connected_aliases(alias: str, graph: dict[str, set[str]]) -> set[str]:
    alias = clean_name(alias)
    if not alias:
        return set()
    if alias not in graph:
        return {alias}
    seen: set[str] = set()
    queue = deque([alias])
    while queue:
        current = queue.popleft()
        if current in seen:
            continue
        seen.add(current)
        for neighbor in graph.get(current, set()):
            if neighbor not in seen:
                queue.append(neighbor)
    return seen


def extract_target_candidates(message: str) -> list[str]:
    text = normalize_text(message)
    lines = [line.strip() for line in str(message).translate(SPACE_TRANSLATION).translate(QUOTE_TRANSLATION).splitlines() if line.strip()]
    candidates: list[str] = []

    for match in re.finditer(r"@([^\s#@]+)", text):
        candidates.append(match.group(1))

    if lines:
        match = re.match(r'^"?([^:："]+)[:：]', lines[0])
        if match:
            candidates.append(match.group(1))

    deduplicated: list[str] = []
    seen = set()
    for candidate in candidates:
        cleaned = clean_name(candidate)
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            deduplicated.append(cleaned)
    return deduplicated


def _name_matches(target: str, sender: str) -> bool:
    target_norm = re.sub(r"\s+", "", target.lower())
    sender_norm = re.sub(r"\s+", "", sender.lower())
    if not target_norm or not sender_norm:
        return False
    if target_norm == sender_norm or target_norm in sender_norm or sender_norm in target_norm:
        return True
    if len(target_norm) >= 4 and len(sender_norm) >= 4 and target_norm[:4] == sender_norm[:4]:
        return True
    return False


def resolve_alias(
    alias: str,
    *,
    main_users: set[str],
    variant_map: dict[str, str],
    graph: dict[str, set[str]],
) -> str | None:
    alias = clean_name(alias)
    if not alias:
        return None
    candidates: set[str] = set()
    for related in connected_aliases(alias, graph):
        for variant in alias_variants(related):
            target = variant_map.get(variant)
            if target:
                candidates.add(target)
    if len(candidates) == 1:
        return next(iter(candidates))
    return None


def resolve_alias_with_fallback(
    alias: str,
    *,
    main_users: set[str],
    variant_map: dict[str, str],
    graph: dict[str, set[str]],
) -> str | None:
    resolved = resolve_alias(alias, main_users=main_users, variant_map=variant_map, graph=graph)
    if resolved:
        return resolved
    cleaned = clean_name(alias)
    if not cleaned:
        return None
    matches = [user for user in main_users if _name_matches(cleaned, user)]
    return matches[0] if len(matches) == 1 else None


def alias_match_keys(alias: str | None, graph: dict[str, set[str]], cache: dict[str, set[str]]) -> set[str]:
    alias = clean_name(alias)
    if not alias:
        return set()
    cached = cache.get(alias)
    if cached is not None:
        return cached
    keys: set[str] = set()
    for related in connected_aliases(alias, graph):
        keys.update(alias_variants(related))
    cache[alias] = keys
    return keys


def load_arm_dataset(
    *,
    path: Path,
    cohort: str,
    group_value: float,
    count_candidates: list[str],
    arm_label: str,
) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header = [str(value) if value is not None else "" for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]

    user_column = pick_header_name(header, USER_COLUMNS, required=True, label="participant-name", path=path)
    age_column = pick_header_name(header, AGE_COLUMNS, required=True, label="age", path=path)
    sex_column = pick_header_name(header, SEX_COLUMNS, required=True, label="sex", path=path)
    bmi_column = pick_header_name(header, BMI_COLUMNS, required=True, label="BMI", path=path)
    count_column = pick_header_name(header, count_candidates, required=True, label="feedback-count", path=path)

    user_idx = header.index(user_column)
    age_idx = header.index(age_column)
    sex_idx = header.index(sex_column)
    bmi_idx = header.index(bmi_column)
    count_idx = header.index(count_column)

    baseline_idx = None
    endline_idx = None
    aux_kg_idx = None
    aux_pct_idx = None
    if cohort == "weight_loss":
        baseline_column = pick_header_name(header, WEIGHT_BASELINE_COLUMNS, required=True, label="baseline-weight", path=path)
        endline_column = pick_header_name(header, WEIGHT_ENDLINE_COLUMNS, required=False, label="endline-weight", path=path)
        aux_kg_column = pick_header_name(header, WEIGHT_LOSS_KG_COLUMNS, required=False, label="weight-loss-kg", path=path)
        aux_pct_column = pick_header_name(header, WEIGHT_LOSS_PCT_COLUMNS, required=False, label="weight-loss-pct", path=path)
        baseline_idx = header.index(baseline_column)
        endline_idx = header.index(endline_column) if endline_column else None
        aux_kg_idx = header.index(aux_kg_column) if aux_kg_column else None
        aux_pct_idx = header.index(aux_pct_column) if aux_pct_column else None
        if endline_idx is None and aux_kg_idx is None:
            raise KeyError(f"{path.name} must include at least one of {WEIGHT_ENDLINE_COLUMNS} or {WEIGHT_LOSS_KG_COLUMNS}.")
    else:
        baseline_column = pick_header_name(header, GLYCEMIC_BASELINE_COLUMNS, required=True, label="baseline-fpg", path=path)
        endline_column = pick_header_name(header, GLYCEMIC_ENDLINE_COLUMNS, required=True, label="endline-fpg", path=path)
        baseline_idx = header.index(baseline_column)
        endline_idx = header.index(endline_column)

    dataset: list[dict[str, Any]] = []
    blank_streak = 0

    for source_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value in (None, "") for value in row):
            blank_streak += 1
            if blank_streak >= 200:
                break
            continue

        raw_user = to_clean_text(row[user_idx])
        user = clean_name(raw_user)
        if not user:
            user = f"MISSING_NICKNAME_{source_row_number}"

        blank_streak = 0
        age = to_float(row[age_idx])
        bmi = to_float(row[bmi_idx])
        sex_f = sex_to_female_indicator(row[sex_idx])
        feedback_count = to_float(row[count_idx]) or 0.0

        record = {
            "arm": arm_label,
            "group": group_value,
            "user": user,
            "user_raw": raw_user,
            "source_row_number": source_row_number,
            "feedback_count": feedback_count,
            "received_any_feedback": 1.0 if feedback_count >= 1 else 0.0,
            "age": age,
            "bmi": bmi,
            "sex_f": sex_f,
        }

        if cohort == "weight_loss":
            baseline_weight = to_float(row[baseline_idx])
            endline_weight = to_float(row[endline_idx]) if endline_idx is not None else None
            weight_loss_kg = to_float(row[aux_kg_idx]) if aux_kg_idx is not None else None
            weight_loss_pct = to_float(row[aux_pct_idx]) if aux_pct_idx is not None else None
            if weight_loss_pct is not None and abs(weight_loss_pct) > 1.0:
                weight_loss_pct = weight_loss_pct / 100.0
            if weight_loss_kg is None and baseline_weight is not None and endline_weight is not None:
                weight_loss_kg = baseline_weight - endline_weight
            if endline_weight is None and baseline_weight is not None and weight_loss_kg is not None:
                endline_weight = baseline_weight - weight_loss_kg
            if weight_loss_pct is None and baseline_weight not in (None, 0, 0.0) and weight_loss_kg is not None:
                weight_loss_pct = weight_loss_kg / baseline_weight
            record.update(
                {
                    "baseline_weight": baseline_weight,
                    "endline_weight": endline_weight,
                    "weight_loss_kg": weight_loss_kg,
                    "weight_loss_pct": weight_loss_pct,
                }
            )
        else:
            baseline_fpg = to_float(row[baseline_idx])
            endline_fpg = to_float(row[endline_idx])
            record.update(
                {
                    "baseline_fpg": baseline_fpg,
                    "endline_fpg": endline_fpg,
                    "fpg_reduction": None if baseline_fpg is None or endline_fpg is None else baseline_fpg - endline_fpg,
                }
            )

        dataset.append(record)

    workbook.close()
    return dataset


def build_diagnostics(
    human_rows: list[dict[str, Any]],
    eps_rows: list[dict[str, Any]],
    human_report: dict[str, Any],
    eps_report: dict[str, Any],
) -> list[dict[str, Any]]:
    diagnostics: list[dict[str, Any]] = []
    for arm_label, rows, report in (
        ("Human", human_rows, human_report),
        ("EPS-human", eps_rows, eps_report),
    ):
        feedback_total = int(round(sum(row["feedback_count"] for row in rows)))
        tagged_total = int(report.get("keyword_message_count", 0))
        reported_feedback_total = int(report.get("resolved_message_count", 0))
        unresolved_total = int(report.get("unresolved_message_count", 0))
        diagnostics.append(
            {
                "arm": arm_label,
                "participant_count": len(rows),
                "feedback_total_from_output_excel": feedback_total,
                "feedback_total_from_report": reported_feedback_total,
                "tagged_message_total_from_report": tagged_total,
                "unresolved_message_count_from_report": unresolved_total,
                "output_matches_report_feedback_total": feedback_total == reported_feedback_total,
            }
        )
    return diagnostics


def build_quality_warnings(diagnostics: list[dict[str, Any]], all_rows: list[dict[str, Any]]) -> list[str]:
    warnings: list[str] = []
    for item in diagnostics:
        arm = item["arm"]
        if not item["output_matches_report_feedback_total"]:
            warnings.append(f"{arm}: feedback totals in the output workbook do not match the JSON linkage report.")

    for arm_label in ("Human", "EPS-human"):
        arm_rows = [row for row in all_rows if row["arm"] == arm_label]
        if not any((row.get("n_quality_events") or 0) > 0 for row in arm_rows):
            warnings.append(f"{arm_label}: no quality-coded feedback events were extracted from the chat history.")
        if not any((row.get("n_latency_events") or 0) > 0 for row in arm_rows):
            warnings.append(f"{arm_label}: no participant-level latency events were extracted from the chat history.")
    return warnings


def extract_feedback_segment(content: str, keyword: str) -> str:
    """Return only the text belonging to the requested feedback hashtag."""
    text = normalize_text(content)
    matches = list(re.finditer(re.escape(keyword), text))
    if not matches:
        return ""
    segment = text[matches[-1].end():]
    next_hash = re.search(r"\s#[^\s#@]+", segment)
    if next_hash:
        segment = segment[: next_hash.start()]
    segment = re.sub(r"^\s*[：:，,。.\-—–]+", "", segment)
    return segment.strip()


def _quality_features(content: str) -> dict[str, float]:
    stripped = re.sub(r"#\S+|@\S+", "", content).strip()
    char_count = len(re.sub(r"\s+", "", stripped))
    encouragement_hits = ENCOURAGEMENT_RE.findall(content)
    return {
        "char_count": float(char_count),
        "has_exercise_params": float(bool(EXERCISE_PARAM_RE.search(content))),
        "has_personalisation": float(bool(PERSONALISE_RE.search(content))),
        "has_advice": float(bool(ADVICE_RE.search(content))),
        "has_encouragement": float(bool(encouragement_hits)),
        "has_goal_plan": float(bool(GOAL_PLAN_RE.search(content))),
        "has_data_reference": float(bool(DATA_REFERENCE_RE.search(content))),
        "positive_count": float(len(encouragement_hits)),
    }


def parse_timestamp(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def keyword_in_content(content: str, keyword: str) -> bool:
    return keyword.lower() in normalize_text(content).lower()


def parse_chat_features(
    *,
    chat_path: Path,
    participants: list[str],
    keyword: str,
    arm_label: str,
) -> dict[str, dict[str, Any]]:
    if not chat_path.exists():
        log(f"{arm_label}: chat file not found ({chat_path.name}); skipping feature extraction.")
        return {}

    main_users = set(participants)
    variant_map = build_variant_map(participants)
    graph = build_alias_graph(chat_path)
    alias_key_cache: dict[str, set[str]] = {}
    participant_match_keys = {user: alias_match_keys(user, graph, alias_key_cache) for user in participants}

    workbook = load_workbook(chat_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header = [str(value) if value is not None else "" for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    content_column = pick_header_name(header, CHAT_CONTENT_COLUMNS, required=True, label="message-content", path=chat_path)
    timestamp_column = pick_header_name(header, CHAT_TIMESTAMP_COLUMNS, required=False, label="timestamp", path=chat_path)
    sender_column = pick_header_name(header, CHAT_SENDER_COLUMNS, required=False, label="sender", path=chat_path)
    nickname_column = pick_header_name(header, CHAT_NICKNAME_COLUMNS, required=False, label="sender-nickname", path=chat_path)

    content_idx = header.index(content_column)
    timestamp_idx = header.index(timestamp_column) if timestamp_column else None
    sender_idx = header.index(sender_column) if sender_column else None
    nickname_idx = header.index(nickname_column) if nickname_column else None

    timeline: list[tuple[datetime | None, str | None, set[str], str]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in row):
            continue
        content = str(row[content_idx]) if row[content_idx] is not None else ""
        timestamp = parse_timestamp(row[timestamp_idx]) if timestamp_idx is not None else None
        sender_raw = clean_name(row[sender_idx]) if sender_idx is not None else ""
        nickname_raw = clean_name(row[nickname_idx]) if nickname_idx is not None else ""
        primary_alias = sender_raw or nickname_raw
        canonical = resolve_alias(primary_alias, main_users=main_users, variant_map=variant_map, graph=graph) if primary_alias else None
        sender_match_keys: set[str] = set()
        for raw_alias in dict.fromkeys([sender_raw, nickname_raw]):
            if raw_alias:
                sender_match_keys.update(alias_match_keys(raw_alias, graph, alias_key_cache))
        timeline.append((timestamp, canonical, sender_match_keys, content))
    workbook.close()

    latencies: dict[str, list[float]] = defaultdict(list)
    qualities: dict[str, list[dict[str, float]]] = defaultdict(list)
    latency_match_modes = {"canonical": 0, "fuzzy": 0}

    for idx, (timestamp_fb, _, _, content) in enumerate(timeline):
        if not keyword_in_content(content, keyword):
            continue
        target: str | None = None
        for candidate in extract_target_candidates(content):
            resolved = resolve_alias_with_fallback(candidate, main_users=main_users, variant_map=variant_map, graph=graph)
            if resolved:
                target = resolved
                break
        if target is None:
            raw_mentions = [clean_name(match.group(1)) for match in re.finditer(r"@([^\s#@]+)", normalize_text(content))]
            for raw_mention in [mention for mention in raw_mentions if mention]:
                for prev_idx2 in range(idx - 1, max(idx - 600, -1), -1):
                    _, prev_canonical, _, _ = timeline[prev_idx2]
                    if prev_canonical and _name_matches(raw_mention, prev_canonical):
                        target = prev_canonical
                        break
                if target:
                    break
        if target is None:
            continue

        feedback_segment = extract_feedback_segment(content, keyword)
        if feedback_segment:
            qualities[target].append(_quality_features(feedback_segment))

        if timestamp_fb is None:
            continue
        target_keys = participant_match_keys.get(target, set())
        look_back = max(idx - 600, 0)
        for prev_idx in range(idx - 1, look_back - 1, -1):
            prev_ts, prev_sender, prev_sender_keys, _ = timeline[prev_idx]
            if prev_ts is None or prev_ts >= timestamp_fb:
                continue
            match_mode: str | None = None
            if prev_sender == target:
                match_mode = "canonical"
            elif target_keys and prev_sender_keys and not target_keys.isdisjoint(prev_sender_keys):
                match_mode = "fuzzy"
            if match_mode:
                latency_minutes = (timestamp_fb - prev_ts).total_seconds() / 60.0
                if 0 < latency_minutes < 1440:
                    latencies[target].append(latency_minutes)
                    latency_match_modes[match_mode] += 1
                break

    result: dict[str, dict[str, Any]] = {}
    for user in participants:
        user_latencies = latencies.get(user, [])
        user_qualities = qualities.get(user, [])
        result[user] = {
            "n_latency_events": len(user_latencies),
            "mean_latency_min": float(np.mean(user_latencies)) if user_latencies else None,
            "n_quality_events": len(user_qualities),
            "mean_char_count": float(np.mean([item["char_count"] for item in user_qualities])) if user_qualities else None,
            "mean_has_exercise_params": float(np.mean([item["has_exercise_params"] for item in user_qualities])) if user_qualities else None,
            "mean_has_personalisation": float(np.mean([item["has_personalisation"] for item in user_qualities])) if user_qualities else None,
            "mean_has_advice": float(np.mean([item["has_advice"] for item in user_qualities])) if user_qualities else None,
            "mean_has_encouragement": float(np.mean([item["has_encouragement"] for item in user_qualities])) if user_qualities else None,
            "mean_has_goal_plan": float(np.mean([item["has_goal_plan"] for item in user_qualities])) if user_qualities else None,
            "mean_has_data_reference": float(np.mean([item["has_data_reference"] for item in user_qualities])) if user_qualities else None,
            "mean_positive_count": float(np.mean([item["positive_count"] for item in user_qualities])) if user_qualities else None,
        }

    log(
        f"{arm_label}: extracted features for {sum(1 for item in result.values() if item['n_quality_events'] > 0)} participants with >=1 quality-coded message; "
        f"latency matches canonical={latency_match_modes['canonical']}, fuzzy={latency_match_modes['fuzzy']}."
    )
    return result


def extract_arm_quality_messages(*, chat_path: Path, keyword: str, arm_label: str) -> list[dict[str, Any]]:
    """Extract one coded row per non-empty feedback hashtag segment."""
    if not chat_path.exists():
        log(f"{arm_label}: chat file not found ({chat_path.name}); skipping message-level content audit.")
        return []

    workbook = load_workbook(chat_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header = [str(value) if value is not None else "" for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    content_column = pick_header_name(header, CHAT_CONTENT_COLUMNS, required=True, label="message-content", path=chat_path)
    content_idx = header.index(content_column)

    coded: list[dict[str, Any]] = []
    keyword_message_count = 0
    empty_segment_count = 0
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value in (None, "") for value in row):
            continue
        content = str(row[content_idx]) if row[content_idx] is not None else ""
        if not keyword_in_content(content, keyword):
            continue
        keyword_message_count += 1
        segment = extract_feedback_segment(content, keyword)
        if not segment:
            empty_segment_count += 1
            continue
        features = _quality_features(segment)
        features.update(
            {
                "arm": arm_label,
                "keyword": keyword,
                "source_row_number": row_number,
            }
        )
        coded.append(features)
    workbook.close()

    log(
        f"{arm_label}: content audit coded {len(coded)} non-empty {keyword} segments "
        f"({keyword_message_count} keyword messages, {empty_segment_count} empty segments)."
    )
    return coded


def _binary_quality_comparison(
    human_messages: list[dict[str, Any]],
    eps_messages: list[dict[str, Any]],
    key: str,
    label: str,
) -> dict[str, Any]:
    human_values = [float(message[key]) for message in human_messages if key in message]
    eps_values = [float(message[key]) for message in eps_messages if key in message]
    n_human = len(human_values)
    n_eps = len(eps_values)
    human_hits = int(sum(human_values))
    eps_hits = int(sum(eps_values))
    human_rate = human_hits / n_human if n_human else None
    eps_rate = eps_hits / n_eps if n_eps else None
    diff = eps_rate - human_rate if human_rate is not None and eps_rate is not None else None

    if diff is not None and n_human > 0 and n_eps > 0:
        se = math.sqrt(eps_rate * (1.0 - eps_rate) / n_eps + human_rate * (1.0 - human_rate) / n_human)
        ci_low = diff - 1.96 * se
        ci_high = diff + 1.96 * se
        pooled = (human_hits + eps_hits) / (n_human + n_eps)
        pooled_se = math.sqrt(pooled * (1.0 - pooled) * (1.0 / n_human + 1.0 / n_eps))
        p_value = normal_two_sided_p(diff / pooled_se) if pooled_se > 0 else (1.0 if abs(diff) < 1e-12 else 0.0)
    else:
        ci_low = ci_high = p_value = None

    return {
        "feature": key,
        "label": label,
        "metric": "message proportion",
        "human_n": n_human,
        "human_hits": human_hits,
        "human_value": human_rate,
        "eps_n": n_eps,
        "eps_hits": eps_hits,
        "eps_value": eps_rate,
        "difference_eps_minus_human": diff,
        "ci_low": ci_low,
        "ci_high": ci_high,
        "p_value": p_value,
        "sig": pstar(p_value),
    }


def _mean_quality_comparison(
    human_messages: list[dict[str, Any]],
    eps_messages: list[dict[str, Any]],
    key: str,
    label: str,
) -> dict[str, Any]:
    human_values = np.array([float(message[key]) for message in human_messages if key in message], dtype=float)
    eps_values = np.array([float(message[key]) for message in eps_messages if key in message], dtype=float)
    n_human = len(human_values)
    n_eps = len(eps_values)
    human_mean = float(np.mean(human_values)) if n_human else None
    eps_mean = float(np.mean(eps_values)) if n_eps else None
    diff = eps_mean - human_mean if human_mean is not None and eps_mean is not None else None

    if diff is not None and n_human > 1 and n_eps > 1:
        human_var = float(np.var(human_values, ddof=1))
        eps_var = float(np.var(eps_values, ddof=1))
        se = math.sqrt(eps_var / n_eps + human_var / n_human)
        ci_low = diff - 1.96 * se
        ci_high = diff + 1.96 * se
        p_value = normal_two_sided_p(diff / se) if se > 0 else (1.0 if abs(diff) < 1e-12 else 0.0)
    else:
        ci_low = ci_high = p_value = None

    return {
        "feature": key,
        "label": label,
        "metric": "message mean",
        "human_n": n_human,
        "human_hits": None,
        "human_value": human_mean,
        "eps_n": n_eps,
        "eps_hits": None,
        "eps_value": eps_mean,
        "difference_eps_minus_human": diff,
        "ci_low": ci_low,
        "ci_high": ci_high,
        "p_value": p_value,
        "sig": pstar(p_value),
    }


def run_keyword_quality_comparison(
    human_messages: list[dict[str, Any]],
    eps_messages: list[dict[str, Any]],
) -> dict[str, Any]:
    rows = [_mean_quality_comparison(human_messages, eps_messages, "char_count", "Segment length (chars)")]
    rows.extend(
        _binary_quality_comparison(human_messages, eps_messages, key, label)
        for key, label in QUALITY_BINARY_FEATURES
    )
    return {
        "module": "Panel C - Message-Level Feedback-Content Audit",
        "unit": "non-empty feedback hashtag segment",
        "human_n_messages": len(human_messages),
        "eps_n_messages": len(eps_messages),
        "features": rows,
        "interpretation": (
            "This is a construct-validity audit, not a formal causal mediation model. "
            "Positive differences mean EPS-human feedback segments have a higher message-level "
            "mean or feature prevalence than Human feedback segments."
        ),
    }


def extract_arm_latencies_all_messages(*, chat_path: Path, keyword: str, arm_label: str) -> list[float]:
    if not chat_path.exists():
        log(f"{arm_label}: chat file not found; skipping arm-level latency extraction.")
        return []

    workbook = load_workbook(chat_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header = [str(value) if value is not None else "" for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    content_column = pick_header_name(header, CHAT_CONTENT_COLUMNS, required=True, label="message-content", path=chat_path)
    timestamp_column = pick_header_name(header, CHAT_TIMESTAMP_COLUMNS, required=False, label="timestamp", path=chat_path)
    sender_column = pick_header_name(header, CHAT_SENDER_COLUMNS, required=False, label="sender", path=chat_path)
    nickname_column = pick_header_name(header, CHAT_NICKNAME_COLUMNS, required=False, label="sender-nickname", path=chat_path)

    content_idx = header.index(content_column)
    timestamp_idx = header.index(timestamp_column) if timestamp_column else None
    sender_idx = header.index(sender_column) if sender_column else None
    nickname_idx = header.index(nickname_column) if nickname_column else None

    timeline: list[tuple[datetime | None, str, str, str]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in row):
            continue
        content = str(row[content_idx]) if row[content_idx] is not None else ""
        timestamp = parse_timestamp(row[timestamp_idx]) if timestamp_idx is not None else None
        sender_name = clean_name(row[sender_idx]) if sender_idx is not None else ""
        sender_nick = clean_name(row[nickname_idx]) if nickname_idx is not None else ""
        timeline.append((timestamp, sender_name, sender_nick, content))
    workbook.close()

    latencies: list[float] = []
    for idx, (timestamp_fb, _, _, content) in enumerate(timeline):
        if timestamp_fb is None or not keyword_in_content(content, keyword):
            continue
        targets = extract_target_candidates(content)
        if not targets:
            continue
        target_name = targets[0]
        look_back = max(idx - 600, 0)
        for prev_idx in range(idx - 1, look_back - 1, -1):
            prev_ts, prev_sender_name, prev_sender_nick, _ = timeline[prev_idx]
            if prev_ts is None or prev_ts >= timestamp_fb:
                continue
            if any(_name_matches(target_name, sender) for sender in (prev_sender_name, prev_sender_nick) if sender):
                latency_minutes = (timestamp_fb - prev_ts).total_seconds() / 60.0
                if 0 < latency_minutes < 1440:
                    latencies.append(latency_minutes)
                break
    log(f"{arm_label}: extracted {len(latencies)} all-message latency values.")
    return latencies


def with_intercept(*columns: np.ndarray) -> np.ndarray:
    n_rows = len(columns[0])
    return np.column_stack([np.ones(n_rows)] + [np.asarray(column, dtype=float) for column in columns])


def fit_ols(y: np.ndarray, design: np.ndarray) -> dict[str, Any]:
    y = np.asarray(y, dtype=float)
    design = np.asarray(design, dtype=float)
    n_obs, n_params = design.shape
    with np.errstate(divide="ignore", invalid="ignore", over="ignore"):
        xtx_inv = np.linalg.pinv(design.T @ design)
        beta = xtx_inv @ design.T @ y
        fitted = design @ beta
        resid = y - fitted

        hat_diag = np.sum(design * (design @ xtx_inv), axis=1)
        hat_diag = np.clip(hat_diag, 0.0, 1.0 - 1e-8)
        hat_denom = np.clip(1.0 - hat_diag, 1e-8, None)
        scale = np.square(resid / hat_denom)
        meat = design.T @ (design * scale[:, None])
        cov_hc3 = xtx_inv @ meat @ xtx_inv
    se = np.sqrt(np.maximum(np.diag(cov_hc3), 0.0))

    t_stat = np.divide(beta, se, out=np.full_like(beta, np.nan), where=se > 0)
    p_value = np.array([normal_two_sided_p(float(value)) for value in t_stat], dtype=float)
    rss = float(np.sum(resid**2))
    tss = float(np.sum((y - np.mean(y)) ** 2))
    r2 = float("nan") if tss <= 0 else 1.0 - rss / tss
    return {"beta": beta, "se": se, "p": p_value, "r2": r2, "n_obs": n_obs}


def rank_average_ties(values: np.ndarray) -> np.ndarray:
    order = np.argsort(values, kind="mergesort")
    ranks = np.empty(len(values), dtype=float)
    sorted_values = values[order]
    start = 0
    while start < len(values):
        end = start + 1
        while end < len(values) and sorted_values[end] == sorted_values[start]:
            end += 1
        ranks[order[start:end]] = (start + 1 + end) / 2.0
        start = end
    return ranks


def pearson_correlation(x: np.ndarray, y: np.ndarray) -> float | None:
    x_centered = x - np.mean(x)
    y_centered = y - np.mean(y)
    denom = math.sqrt(float(np.sum(x_centered**2) * np.sum(y_centered**2)))
    if denom <= 0:
        return None
    return float(np.sum(x_centered * y_centered) / denom)


def spearman_rho(x: np.ndarray, y: np.ndarray) -> dict[str, float | None]:
    if len(x) < 3 or np.all(x == x[0]) or np.all(y == y[0]):
        return {"rho": None, "p": None}
    x_rank = rank_average_ties(x)
    y_rank = rank_average_ties(y)
    rho = pearson_correlation(x_rank, y_rank)
    if rho is None:
        return {"rho": None, "p": None}
    if abs(rho) >= 1.0:
        return {"rho": float(rho), "p": 0.0}
    t_like = rho * math.sqrt((len(x) - 2.0) / max(1e-12, 1.0 - rho**2))
    return {"rho": float(rho), "p": normal_two_sided_p(t_like)}


def fit_linear_slope(x: np.ndarray, y: np.ndarray) -> dict[str, float | None]:
    if len(x) < 3 or np.all(x == x[0]):
        return {"slope": None, "intercept": None}
    x_mean = float(np.mean(x))
    y_mean = float(np.mean(y))
    ss_x = float(np.sum((x - x_mean) ** 2))
    if ss_x <= 0:
        return {"slope": None, "intercept": None}
    slope = float(np.sum((x - x_mean) * (y - y_mean)) / ss_x)
    return {"slope": slope, "intercept": float(y_mean - slope * x_mean)}


def mann_whitney_u(x: list[float], y: list[float]) -> dict[str, Any]:
    n1, n2 = len(x), len(y)
    if n1 < 3 or n2 < 3:
        return {"u": None, "z": None, "p": None, "n1": n1, "n2": n2}

    combined = sorted(enumerate(list(x) + list(y)), key=lambda item: item[1])
    ranks = [0.0] * (n1 + n2)
    index = 0
    while index < len(combined):
        end = index
        while end < len(combined) - 1 and combined[end + 1][1] == combined[index][1]:
            end += 1
        average_rank = (index + end + 2) / 2.0
        for rank_idx in range(index, end + 1):
            ranks[combined[rank_idx][0]] = average_rank
        index = end + 1

    r1 = sum(ranks[:n1])
    u1 = r1 - n1 * (n1 + 1) / 2.0
    u2 = n1 * n2 - u1
    u = min(u1, u2)
    mu = n1 * n2 / 2.0
    sigma = math.sqrt(n1 * n2 * (n1 + n2 + 1) / 12.0)
    if sigma <= 0:
        return {"u": float(u), "z": None, "p": None, "n1": n1, "n2": n2}
    z = (u - mu) / sigma
    return {"u": float(u), "z": float(z), "p": float(2.0 * (1.0 - STANDARD_NORMAL.cdf(abs(z)))), "n1": n1, "n2": n2}


def arm_latency_descriptives(latencies: list[float], label: str) -> dict[str, Any]:
    if not latencies:
        return {"arm": label, "n": 0, "mean": None, "median": None, "sd": None, "p25": None, "p75": None}
    values = np.array(latencies, dtype=float)
    return {
        "arm": label,
        "n": len(values),
        "mean": float(np.mean(values)),
        "median": float(np.median(values)),
        "sd": float(np.std(values, ddof=1)) if len(values) > 1 else None,
        "p25": float(np.percentile(values, 25)),
        "p75": float(np.percentile(values, 75)),
    }


def required_keys(spec: OutcomeSpec) -> list[str]:
    return [spec.key, "group", "feedback_count", spec.covariate_key, "age", "bmi", "sex_f"]


def covariate_matrix(rows: list[dict[str, Any]], spec: OutcomeSpec) -> np.ndarray:
    return np.array([[row[spec.covariate_key], row["age"], row["bmi"], row["sex_f"]] for row in rows], dtype=float)


def run_interaction_model(rows: list[dict[str, Any]], spec: OutcomeSpec) -> dict[str, Any]:
    valid = [row for row in rows if all(row.get(key) is not None for key in required_keys(spec))]
    if len(valid) < 12:
        return {"error": f"Too few complete cases ({len(valid)}) for interaction model.", "outcome": spec.key, "outcome_title": spec.title}

    groups = np.array([row["group"] for row in valid], dtype=float)
    counts = np.array([row["feedback_count"] for row in valid], dtype=float)
    outcome = np.array([row[spec.key] for row in valid], dtype=float)
    covars = covariate_matrix(valid, spec)

    human_mask = groups == HUMAN_GROUP
    eps_mask = groups == EPS_GROUP
    centre = float(np.mean(counts[human_mask])) if np.any(human_mask) else float(np.mean(counts))
    centred_counts = counts - centre
    interaction = groups * centred_counts

    fit = fit_ols(outcome, with_intercept(groups, centred_counts, interaction, covars))
    human_corr = spearman_rho(counts[human_mask], outcome[human_mask])
    eps_corr = spearman_rho(counts[eps_mask], outcome[eps_mask])
    return {
        "module": "Module 1 - Interaction / Dose-Response Model",
        "outcome": spec.key,
        "outcome_title": spec.title,
        "n_total": len(valid),
        "n_human": int(np.sum(human_mask)),
        "n_eps": int(np.sum(eps_mask)),
        "feedback_count_centre": centre,
        "arm_effect_at_centre": {
            "label": "beta_1 (arm effect at equal feedback count)",
            "estimate": float(fit["beta"][1]),
            "se": float(fit["se"][1]),
            "p_value": float(fit["p"][1]),
            "sig": pstar(float(fit["p"][1])),
        },
        "dose_response_slope_human": {
            "label": "beta_2 (dose-response slope in Human arm)",
            "estimate": float(fit["beta"][2]),
            "se": float(fit["se"][2]),
            "p_value": float(fit["p"][2]),
            "sig": pstar(float(fit["p"][2])),
        },
        "interaction_term": {
            "label": "beta_3 (arm x feedback_count interaction)",
            "estimate": float(fit["beta"][3]),
            "se": float(fit["se"][3]),
            "p_value": float(fit["p"][3]),
            "sig": pstar(float(fit["p"][3])),
        },
        "r2": float(fit["r2"]),
        "interpretation": "A persistent arm effect at equal feedback frequency indicates that the outcome difference is not explained by feedback count alone.",
        "within_arm_dose_response": {
            "human_spearman_rho": human_corr["rho"],
            "human_spearman_p": human_corr["p"],
            "eps_spearman_rho": eps_corr["rho"],
            "eps_spearman_p": eps_corr["p"],
        },
    }


def greedy_match(eps_counts: np.ndarray, human_counts: np.ndarray, caliper: float | None) -> list[tuple[int, int]]:
    available = list(range(len(human_counts)))
    pairs: list[tuple[int, int]] = []
    for eps_idx in np.argsort(eps_counts):
        if not available:
            break
        best_dist, best_human_idx = min((abs(eps_counts[eps_idx] - human_counts[human_idx]), human_idx) for human_idx in available)
        if caliper is not None and best_dist > caliper:
            continue
        pairs.append((int(eps_idx), int(best_human_idx)))
        available.remove(best_human_idx)
    return pairs


def run_matching_analysis(rows: list[dict[str, Any]], spec: OutcomeSpec) -> dict[str, Any]:
    valid = [row for row in rows if all(row.get(key) is not None for key in required_keys(spec))]
    if len(valid) < 12:
        return {"error": f"Too few complete cases ({len(valid)}) for matching.", "outcome": spec.key, "outcome_title": spec.title}

    human_rows = [row for row in valid if row["group"] == HUMAN_GROUP]
    eps_rows = [row for row in valid if row["group"] == EPS_GROUP]
    if not human_rows or not eps_rows:
        return {"error": "One arm is empty after filtering.", "outcome": spec.key, "outcome_title": spec.title}

    human_counts = np.array([row["feedback_count"] for row in human_rows], dtype=float)
    eps_counts = np.array([row["feedback_count"] for row in eps_rows], dtype=float)
    overlap_lo = max(float(np.min(human_counts)), float(np.min(eps_counts)))
    overlap_hi = min(float(np.max(human_counts)), float(np.max(eps_counts)))

    human_in = [row for row in human_rows if overlap_lo <= row["feedback_count"] <= overlap_hi]
    eps_in = [row for row in eps_rows if overlap_lo <= row["feedback_count"] <= overlap_hi]
    if len(human_in) < 3 or len(eps_in) < 3:
        return {"error": "Insufficient overlap region for matching.", "outcome": spec.key, "outcome_title": spec.title, "overlap_lo": overlap_lo, "overlap_hi": overlap_hi}

    human_in_counts = np.array([row["feedback_count"] for row in human_in], dtype=float)
    eps_in_counts = np.array([row["feedback_count"] for row in eps_in], dtype=float)
    pooled = np.concatenate([human_in_counts, eps_in_counts])
    caliper = float(0.5 * np.std(pooled, ddof=1)) if len(pooled) > 1 else None
    pairs = greedy_match(eps_in_counts, human_in_counts, caliper)
    if len(pairs) < 4:
        return {"error": f"Only {len(pairs)} matched pairs were available.", "outcome": spec.key, "outcome_title": spec.title}

    matched_eps = [eps_in[eps_idx] for eps_idx, _ in pairs]
    matched_human = [human_in[human_idx] for _, human_idx in pairs]
    matched_all = matched_eps + matched_human

    groups = np.array([row["group"] for row in matched_all], dtype=float)
    outcome = np.array([row[spec.key] for row in matched_all], dtype=float)
    counts = np.array([row["feedback_count"] for row in matched_all], dtype=float)
    covars = covariate_matrix(matched_all, spec)
    fit_adj = fit_ols(outcome, with_intercept(groups, covars))
    fit_count_adj = fit_ols(outcome, with_intercept(groups, counts, covars))

    return {
        "module": "Module 2 - Nearest-Neighbour Matching",
        "outcome": spec.key,
        "outcome_title": spec.title,
        "overlap_lo": overlap_lo,
        "overlap_hi": overlap_hi,
        "n_human_in_overlap": len(human_in),
        "n_eps_in_overlap": len(eps_in),
        "n_matched_pairs": len(pairs),
        "caliper": caliper,
        "matched_feedback_balance": {
            "mean_eps": float(np.mean([eps_in[eps_idx]["feedback_count"] for eps_idx, _ in pairs])),
            "mean_human": float(np.mean([human_in[human_idx]["feedback_count"] for _, human_idx in pairs])),
        },
        "unadjusted_arm_diff": float(np.mean([row[spec.key] for row in matched_eps]) - np.mean([row[spec.key] for row in matched_human])),
        "covariate_adjusted": {
            "label": "Arm effect (EPS - Human) in matched subsample",
            "estimate": float(fit_adj["beta"][1]),
            "se": float(fit_adj["se"][1]),
            "p_value": float(fit_adj["p"][1]),
            "sig": pstar(float(fit_adj["p"][1])),
        },
        "count_and_covariate_adjusted": {
            "label": "Arm effect after further controlling for feedback count",
            "estimate": float(fit_count_adj["beta"][1]),
            "se": float(fit_count_adj["se"][1]),
            "p_value": float(fit_count_adj["p"][1]),
            "sig": pstar(float(fit_count_adj["p"][1])),
        },
        "interpretation": "A persistent arm effect in the matched subsample indicates that the outcome difference is not explained by feedback count alone.",
    }


def describe_features(rows: list[dict[str, Any]]) -> dict[str, Any]:
    def arm_stats(arm_rows: list[dict[str, Any]], key: str) -> dict[str, Any]:
        values = [row[key] for row in arm_rows if row.get(key) is not None]
        if not values:
            return {"n": 0, "mean": None, "sd": None, "median": None}
        array = np.array(values, dtype=float)
        return {
            "n": len(array),
            "mean": float(np.mean(array)),
            "sd": float(np.std(array, ddof=1)) if len(array) > 1 else None,
            "median": float(np.median(array)),
        }

    keys = [
        "feedback_count",
        "mean_latency_min",
        "mean_char_count",
        "mean_has_exercise_params",
        "mean_has_personalisation",
        "mean_has_advice",
        "mean_has_encouragement",
        "mean_has_goal_plan",
        "mean_has_data_reference",
        "mean_positive_count",
    ]
    human_rows = [row for row in rows if row["group"] == HUMAN_GROUP]
    eps_rows = [row for row in rows if row["group"] == EPS_GROUP]
    return {key: {"human": arm_stats(human_rows, key), "eps": arm_stats(eps_rows, key)} for key in keys}


def maybe_build_plot(
    *,
    rows: list[dict[str, Any]],
    config: CohortConfig,
    quality_comparison: dict[str, Any],
    human_latencies: list[float],
    eps_latencies: list[float],
    plot_path: Path,
) -> str:
    try:
        with contextlib.redirect_stderr(io.StringIO()), contextlib.redirect_stdout(io.StringIO()):
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.gridspec as gridspec
            import matplotlib.pyplot as plt
    except Exception as exc:
        return f"Skipped plot generation because matplotlib is unavailable: {exc}"

    colors = {"Human": "#2563EB", "EPS-human": "#EA580C"}
    fig = plt.figure(figsize=(13, 9), facecolor="white")
    gs = gridspec.GridSpec(2, 2, figure=fig, hspace=0.42, wspace=0.32)

    human_rows = [row for row in rows if row["group"] == HUMAN_GROUP]
    eps_rows = [row for row in rows if row["group"] == EPS_GROUP]

    ax_a = fig.add_subplot(gs[0, 0])
    for arm_rows, label in ((human_rows, "Human"), (eps_rows, "EPS-human")):
        xs = np.array([row["feedback_count"] for row in arm_rows if row.get(config.primary_plot_outcome) is not None], dtype=float)
        ys = np.array([row[config.primary_plot_outcome] for row in arm_rows if row.get(config.primary_plot_outcome) is not None], dtype=float)
        if len(xs) == 0:
            continue
        ax_a.scatter(xs, ys, label=label, color=colors[label], alpha=0.65, s=28)
        slope_info = fit_linear_slope(xs, ys)
        if slope_info["slope"] is not None and slope_info["intercept"] is not None:
            x_grid = np.linspace(float(np.min(xs)), float(np.max(xs)), 100)
            y_grid = slope_info["slope"] * x_grid + slope_info["intercept"]
            ax_a.plot(x_grid, y_grid, color=colors[label], linewidth=1.8)
    ax_a.set_xlabel("Feedback count")
    ax_a.set_ylabel(config.plot_ylabel)
    ax_a.set_title("A. Within-arm dose-response")
    ax_a.legend(fontsize=8)
    ax_a.spines[["top", "right"]].set_visible(False)

    ax_b = fig.add_subplot(gs[0, 1])
    for arm_rows, label in ((human_rows, "Human"), (eps_rows, "EPS-human")):
        values = [row["feedback_count"] for row in arm_rows]
        if values:
            ax_b.hist(values, bins=20, density=True, alpha=0.55, color=colors[label], label=label)
    ax_b.set_xlabel("Feedback count")
    ax_b.set_ylabel("Density")
    ax_b.set_title("B. Feedback frequency distribution")
    ax_b.legend(fontsize=8)
    ax_b.spines[["top", "right"]].set_visible(False)

    ax_c = fig.add_subplot(gs[1, 0])
    binary_features = [item for item in quality_comparison.get("features", []) if item.get("metric") == "message proportion"]
    if binary_features:
        labels = [item["label"] for item in binary_features]
        human_values = [100.0 * item["human_value"] if item["human_value"] is not None else np.nan for item in binary_features]
        eps_values = [100.0 * item["eps_value"] if item["eps_value"] is not None else np.nan for item in binary_features]
        positions = np.arange(len(labels))
        width = 0.38
        ax_c.bar(positions - width / 2, human_values, width, color=colors["Human"], alpha=0.78, label="Human")
        ax_c.bar(positions + width / 2, eps_values, width, color=colors["EPS-human"], alpha=0.78, label="EPS-human")
        ax_c.set_xticks(positions)
        ax_c.set_xticklabels(labels, rotation=28, ha="right")
        ax_c.set_ylim(0, 105)
        ax_c.legend(fontsize=8)
    else:
        ax_c.text(0.5, 0.5, "No coded messages", ha="center", va="center", transform=ax_c.transAxes)
    ax_c.set_ylabel("Messages with feature (%)")
    ax_c.set_title("C. Content audit")
    ax_c.spines[["top", "right"]].set_visible(False)

    ax_d = fig.add_subplot(gs[1, 1])
    for values, label in ((human_latencies, "Human"), (eps_latencies, "EPS-human")):
        if values:
            ax_d.hist(values, bins=25, density=True, alpha=0.55, color=colors[label], label=label)
            ax_d.axvline(float(np.median(values)), color=colors[label], linewidth=1.7)
    if human_latencies or eps_latencies:
        ax_d.legend(fontsize=8)
    else:
        ax_d.text(0.5, 0.5, "No latency matches", ha="center", va="center", transform=ax_d.transAxes)
    ax_d.set_xlabel("Response latency (min)")
    ax_d.set_ylabel("Density")
    ax_d.set_title("D. Response latency")
    ax_d.spines[["top", "right"]].set_visible(False)

    fig.suptitle(config.title, fontsize=13, fontweight="bold")
    ensure_parent_dir(plot_path)
    fig.savefig(plot_path, dpi=160, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return f"Saved plot to {plot_path.name}"


def write_cell(
    ws,
    row: int,
    col: int,
    value: Any,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    align: Alignment | None = None,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    cell.border = CELL_BORDER


def autosize_columns(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for idx, width in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(width + 3, 60)


def write_section_header(ws, row: int, text: str, n_cols: int = 2) -> None:
    write_cell(ws, row, 1, text, font=SECTION_FONT, fill=SECTION_FILL, align=LEFT)
    for col in range(2, n_cols + 1):
        write_cell(ws, row, col, None, fill=SECTION_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)


def write_kv_rows(ws, start_row: int, items: list[tuple[str, Any]]) -> int:
    row = start_row
    for label, value in items:
        write_cell(ws, row, 1, label, align=LEFT)
        write_cell(ws, row, 2, value, align=CENTER)
        row += 1
    return row


def dataset_headers_for_cohort(cohort: str) -> list[str]:
    base = [
        "arm",
        "group",
        "user",
        "source_row_number",
        "feedback_count",
        "received_any_feedback",
        "age",
        "bmi",
        "sex_f",
    ]
    if cohort == "weight_loss":
        clinical = ["baseline_weight", "endline_weight", "weight_loss_kg", "weight_loss_pct"]
    else:
        clinical = ["baseline_fpg", "endline_fpg", "fpg_reduction"]
    feature = [
        "n_latency_events",
        "mean_latency_min",
        "n_quality_events",
        "mean_char_count",
        "mean_has_exercise_params",
        "mean_has_personalisation",
        "mean_has_advice",
        "mean_has_encouragement",
        "mean_has_goal_plan",
        "mean_has_data_reference",
        "mean_positive_count",
    ]
    return base + clinical + feature


def module_1_rows(result: dict[str, Any]) -> list[tuple[str, Any]]:
    return [
        ("n_human", result["n_human"]),
        ("n_eps", result["n_eps"]),
        ("feedback_count_centre", format_number(result["feedback_count_centre"], 3)),
        ("beta_1 estimate", format_number(result["arm_effect_at_centre"]["estimate"])),
        ("beta_1 SE", format_number(result["arm_effect_at_centre"]["se"])),
        ("beta_1 p-value", format_number(result["arm_effect_at_centre"]["p_value"])),
        ("beta_1 sig", result["arm_effect_at_centre"]["sig"]),
        ("beta_2 estimate", format_number(result["dose_response_slope_human"]["estimate"])),
        ("beta_2 p-value", format_number(result["dose_response_slope_human"]["p_value"])),
        ("beta_2 sig", result["dose_response_slope_human"]["sig"]),
        ("beta_3 estimate", format_number(result["interaction_term"]["estimate"])),
        ("beta_3 p-value", format_number(result["interaction_term"]["p_value"])),
        ("beta_3 sig", result["interaction_term"]["sig"]),
        ("R2", format_number(result["r2"])),
        ("Human Spearman rho", format_number(result["within_arm_dose_response"]["human_spearman_rho"])),
        ("Human Spearman p", format_number(result["within_arm_dose_response"]["human_spearman_p"])),
        ("EPS Spearman rho", format_number(result["within_arm_dose_response"]["eps_spearman_rho"])),
        ("EPS Spearman p", format_number(result["within_arm_dose_response"]["eps_spearman_p"])),
        ("Interpretation", result["interpretation"]),
    ]


def module_2_rows(result: dict[str, Any]) -> list[tuple[str, Any]]:
    return [
        ("Overlap [lo, hi]", f"[{format_number(result['overlap_lo'], 2)}, {format_number(result['overlap_hi'], 2)}]"),
        ("n Human in overlap", result["n_human_in_overlap"]),
        ("n EPS in overlap", result["n_eps_in_overlap"]),
        ("Matched pairs", result["n_matched_pairs"]),
        ("Caliper", format_number(result["caliper"], 3)),
        ("Mean count EPS (matched)", format_number(result["matched_feedback_balance"]["mean_eps"], 2)),
        ("Mean count Human (matched)", format_number(result["matched_feedback_balance"]["mean_human"], 2)),
        ("Raw arm diff", format_number(result["unadjusted_arm_diff"])),
        ("Covariate-adjusted estimate", format_number(result["covariate_adjusted"]["estimate"])),
        ("Covariate-adjusted SE", format_number(result["covariate_adjusted"]["se"])),
        ("Covariate-adjusted p", format_number(result["covariate_adjusted"]["p_value"])),
        ("Covariate-adjusted sig", result["covariate_adjusted"]["sig"]),
        ("Count+covariate estimate", format_number(result["count_and_covariate_adjusted"]["estimate"])),
        ("Count+covariate p", format_number(result["count_and_covariate_adjusted"]["p_value"])),
        ("Count+covariate sig", result["count_and_covariate_adjusted"]["sig"]),
        ("Interpretation", result["interpretation"]),
    ]


def write_module_sheet(workbook: Workbook, *, sheet_name: str, results: list[dict[str, Any]], row_builder) -> None:
    ws = workbook.create_sheet(sheet_name)
    for col_idx, header in enumerate(["Parameter", "Value"], start=1):
        write_cell(ws, 1, col_idx, header, font=HEADER_FONT, fill=LINE_FILL, align=CENTER)
    row = 2
    for result in results:
        if "error" in result:
            write_section_header(ws, row, f"{result.get('outcome', '?')} - SKIPPED: {result['error']}")
            row += 1
            continue
        write_section_header(ws, row, f"Outcome: {result['outcome']} (n={result.get('n_total', result.get('n_obs', '?'))})")
        row += 1
        row = write_kv_rows(ws, row, row_builder(result))
        row += 1
    autosize_columns(ws)


def format_quality_value(item: dict[str, Any], value: float | None) -> str:
    if item.get("metric") == "message proportion":
        return format_pct(value)
    return format_number(value, 1)


def write_content_audit_sheet(workbook: Workbook, *, comparison: dict[str, Any]) -> None:
    ws = workbook.create_sheet("Panel C - Content Audit")
    headers = ["Feature", "Human", "EPS-human", "Difference", "95% CI", "P value", "Metric"]
    for col_idx, header in enumerate(headers, start=1):
        write_cell(ws, 1, col_idx, header, font=HEADER_FONT, fill=LINE_FILL, align=CENTER)
    row = 2
    write_section_header(
        ws,
        row,
        f"Message-level coded segments: Human n={comparison.get('human_n_messages', 0)}, EPS-human n={comparison.get('eps_n_messages', 0)}",
        len(headers),
    )
    row += 1
    for item in comparison.get("features", []):
        ci_text = f"[{format_quality_value(item, item.get('ci_low'))}, {format_quality_value(item, item.get('ci_high'))}]"
        values = [
            item["label"],
            format_quality_value(item, item.get("human_value")),
            format_quality_value(item, item.get("eps_value")),
            format_quality_value(item, item.get("difference_eps_minus_human")),
            ci_text,
            format_number(item.get("p_value")),
            item.get("metric"),
        ]
        for col_idx, value in enumerate(values, start=1):
            write_cell(ws, row, col_idx, value, align=LEFT if col_idx in (1, 7) else CENTER)
        row += 1
    row += 1
    write_section_header(ws, row, "Interpretation", len(headers))
    row += 1
    write_cell(ws, row, 1, comparison.get("interpretation"), align=LEFT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    autosize_columns(ws)


def write_latency_descriptive_sheet(workbook: Workbook, *, arm_stats: dict[str, Any]) -> None:
    ws = workbook.create_sheet("Panel D - Latency")
    headers = ["Arm", "n messages", "Mean latency (min)", "Median latency (min)", "IQR [P25, P75]", "SD latency (min)"]
    for col_idx, header in enumerate(headers, start=1):
        write_cell(ws, 1, col_idx, header, font=HEADER_FONT, fill=LINE_FILL, align=CENTER)
    for row_idx, key in enumerate(["eps", "human"], start=2):
        item = arm_stats.get(key, {})
        values = [
            item.get("arm"),
            item.get("n"),
            format_number(item.get("mean"), 2),
            format_number(item.get("median"), 2),
            f"[{format_number(item.get('p25'), 1)}, {format_number(item.get('p75'), 1)}]",
            format_number(item.get("sd"), 2),
        ]
        for col_idx, value in enumerate(values, start=1):
            write_cell(ws, row_idx, col_idx, value, align=LEFT if col_idx == 1 else CENTER)
    row = 5
    write_section_header(ws, row, "Mann-Whitney comparison", len(headers))
    row += 1
    mw = arm_stats.get("mann_whitney", {})
    values = [
        ("U", format_number(mw.get("u"), 1)),
        ("z", format_number(mw.get("z"))),
        ("P value", format_number(mw.get("p"))),
        ("Significance", pstar(mw.get("p")) if mw.get("p") is not None else "NA"),
    ]
    row = write_kv_rows(ws, row, values)
    write_section_header(ws, row + 1, "Definition", len(headers))
    write_cell(
        ws,
        row + 2,
        1,
        "Latency is the elapsed time between the participant's most recent exercise-record post and the delivery of the exercise-feedback reply within a 24-hour search window.",
        align=LEFT,
    )
    ws.merge_cells(start_row=row + 2, start_column=1, end_row=row + 2, end_column=len(headers))
    autosize_columns(ws)


def write_readme_sheet(workbook: Workbook, *, config: CohortConfig, human_keyword: str, eps_keyword: str, n_boot: int, seed: int) -> None:
    ws = workbook.create_sheet("README", 0)
    lines = [
        (config.title, HEADER_FONT, LINE_FILL),
        ("", None, None),
        ("This workbook reports the updated article logic: frequency-control analyses, a message-level content audit, and response-latency descriptives.", None, None),
        ("The content audit is construct-validity evidence for individualized exercise prescription; it is not a formal causal mediation analysis.", None, None),
        ("", None, None),
        ("Inputs", SECTION_FONT, SECTION_FILL),
        (f"Human keyword: {human_keyword}", None, None),
        (f"EPS keyword: {eps_keyword}", None, None),
        (f"Random seed: {seed}", None, None),
        (f"Deprecated bootstrap argument retained for CLI compatibility: {n_boot} (not used in this workflow)", None, None),
        ("", None, None),
        ("Panel A - Interaction / dose-response model", SECTION_FONT, SECTION_FILL),
        ("  Fits outcome ~ arm + centred feedback count + arm x count + covariates.", None, None),
        ("  The arm effect at equal feedback count tests whether the arm difference persists after frequency control.", None, None),
        ("  The arm x count term tests whether the count-outcome slope differs by arm.", None, None),
        ("", None, None),
        ("Panel B - Nearest-neighbour matching", SECTION_FONT, SECTION_FILL),
        ("  Greedy 1:1 nearest-neighbour matching on feedback count within the overlap region.", None, None),
        ("  The matched model estimates the arm effect after approximate frequency balance.", None, None),
        ("", None, None),
        ("Panel C - Message-level content audit", SECTION_FONT, SECTION_FILL),
        ("  Codes feedback segment length and six individualized-exercise-prescription content features.", None, None),
        ("  Between-arm differences in proportions use two-sided z tests; segment length uses Welch-style SE.", None, None),
        ("", None, None),
        ("Panel D - Response latency", SECTION_FONT, SECTION_FILL),
        ("  Computes elapsed time from the participant's most recent exercise-record post to the feedback reply.", None, None),
        ("  Reports arm-level medians, IQRs, and a Mann-Whitney comparison.", None, None),
        ("", None, None),
        ("Notes", SECTION_FONT, SECTION_FILL),
        ("  All OLS standard errors are HC3 heteroskedasticity-robust.", None, None),
        ("  Diagnostics still compare the checkin workbook counts against the linkage JSON reports.", None, None),
    ]
    for row_idx, (text, font, fill) in enumerate(lines, start=1):
        write_cell(ws, row_idx, 1, text, font=font, fill=fill, align=LEFT)
    ws.column_dimensions["A"].width = 96


def write_diagnostics_sheet(
    workbook: Workbook,
    *,
    diagnostics: list[dict[str, Any]],
    warnings: list[str],
    plot_status: str,
) -> None:
    ws = workbook.create_sheet("Diagnostics")
    headers = [
        "arm",
        "participant_count",
        "feedback_total_from_output_excel",
        "feedback_total_from_report",
        "tagged_message_total_from_report",
        "unresolved_message_count_from_report",
        "output_matches_report_feedback_total",
    ]
    for col_idx, header in enumerate(headers, start=1):
        write_cell(ws, 1, col_idx, header, font=HEADER_FONT, fill=LINE_FILL, align=CENTER)
    for row_idx, item in enumerate(diagnostics, start=2):
        for col_idx, header in enumerate(headers, start=1):
            write_cell(ws, row_idx, col_idx, item.get(header), align=CENTER)

    row = len(diagnostics) + 3
    write_section_header(ws, row, "Warnings", 2)
    row += 1
    warning_items = warnings or ["No warnings detected."]
    for message in warning_items:
        write_cell(ws, row, 1, message, fill=WARN_FILL if warnings else None, align=LEFT)
        row += 1
    row += 1
    write_section_header(ws, row, "Plot status", 2)
    row += 1
    write_cell(ws, row, 1, plot_status, align=LEFT)
    autosize_columns(ws)


def write_dataset_sheet(workbook: Workbook, *, cohort: str, rows: list[dict[str, Any]]) -> None:
    ws = workbook.create_sheet("Participant Dataset")
    headers = dataset_headers_for_cohort(cohort)
    for col_idx, header in enumerate(headers, start=1):
        write_cell(ws, 1, col_idx, header, font=HEADER_FONT, fill=LINE_FILL, align=CENTER)
    for row_idx, item in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            write_cell(ws, row_idx, col_idx, item.get(header), align=CENTER)
    autosize_columns(ws)


def write_descriptives_sheet(workbook: Workbook, *, desc: dict[str, Any]) -> None:
    ws = workbook.create_sheet("Descriptive Statistics")
    headers = ["Feature", "Human n", "Human mean", "Human SD", "EPS n", "EPS mean", "EPS SD"]
    for col_idx, header in enumerate(headers, start=1):
        write_cell(ws, 1, col_idx, header, font=HEADER_FONT, fill=LINE_FILL, align=CENTER)
    row = 2
    for feature, values in desc.items():
        human = values["human"]
        eps = values["eps"]
        entries = [
            feature,
            human["n"],
            round(human["mean"], 4) if human["mean"] is not None else "NA",
            round(human["sd"], 4) if human["sd"] is not None else "NA",
            eps["n"],
            round(eps["mean"], 4) if eps["mean"] is not None else "NA",
            round(eps["sd"], 4) if eps["sd"] is not None else "NA",
        ]
        for col_idx, value in enumerate(entries, start=1):
            write_cell(ws, row, col_idx, value, align=CENTER if col_idx > 1 else LEFT)
        row += 1
    autosize_columns(ws)


def write_results_workbook(
    *,
    config: CohortConfig,
    diagnostics: list[dict[str, Any]],
    warnings: list[str],
    rows: list[dict[str, Any]],
    desc: dict[str, Any],
    m1_results: list[dict[str, Any]],
    m2_results: list[dict[str, Any]],
    quality_comparison: dict[str, Any],
    arm_latency_stats: dict[str, Any],
    human_keyword: str,
    eps_keyword: str,
    n_boot: int,
    seed: int,
    plot_status: str,
    output_xlsx: Path,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    write_readme_sheet(workbook, config=config, human_keyword=human_keyword, eps_keyword=eps_keyword, n_boot=n_boot, seed=seed)
    write_diagnostics_sheet(workbook, diagnostics=diagnostics, warnings=warnings, plot_status=plot_status)
    write_dataset_sheet(workbook, cohort=config.key, rows=rows)
    write_descriptives_sheet(workbook, desc=desc)
    write_module_sheet(workbook, sheet_name="Panel A - Interaction", results=m1_results, row_builder=module_1_rows)
    write_module_sheet(workbook, sheet_name="Panel B - Matching", results=m2_results, row_builder=module_2_rows)
    write_content_audit_sheet(workbook, comparison=quality_comparison)
    write_latency_descriptive_sheet(workbook, arm_stats=arm_latency_stats)
    ensure_parent_dir(output_xlsx)
    workbook.save(output_xlsx)


def build_markdown_report(
    *,
    config: CohortConfig,
    diagnostics: list[dict[str, Any]],
    warnings: list[str],
    desc: dict[str, Any],
    m1_results: list[dict[str, Any]],
    m2_results: list[dict[str, Any]],
    quality_comparison: dict[str, Any],
    arm_latency_stats: dict[str, Any],
    plot_status: str,
) -> str:
    lines = [
        f"# {config.title}",
        "",
        f"*Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}*",
        "",
        "## Scope",
        "",
        f"This workflow augments the tagged-checkin linkage analysis in the {config.label} with frequency-control analyses, a message-level content audit, and response-latency descriptives.",
        "",
        "The message-level audit is construct-validity evidence for individualized exercise prescription. It is not a formal causal mediation analysis.",
        "",
        "## Diagnostics",
        "",
    ]
    for item in diagnostics:
        lines.append(
            f"- {item['arm']}: workbook feedback total={item['feedback_total_from_output_excel']}, report feedback total={item['feedback_total_from_report']}, "
            f"tagged messages in chat export={item['tagged_message_total_from_report']}, unresolved tagged messages={item['unresolved_message_count_from_report']}."
        )
    lines.extend(["", "## Warnings", ""])
    for warning in warnings or ["No warnings detected."]:
        lines.append(f"- {warning}")

    def append_module(title: str, results: list[dict[str, Any]], summary_fn) -> None:
        lines.extend(["", f"## {title}", ""])
        for result in results:
            if "error" in result:
                lines.append(f"- `{result.get('outcome', '?')}` skipped: {result['error']}")
                continue
            lines.append(f"### `{result['outcome']}`")
            for sentence in summary_fn(result):
                lines.append(f"- {sentence}")
            lines.append("")

    append_module(
        "Panel A - Interaction / Dose-Response",
        m1_results,
        lambda result: [
            f"beta_1 (arm effect at equal feedback count) = {format_number(result['arm_effect_at_centre']['estimate'])} (SE {format_number(result['arm_effect_at_centre']['se'])}, p={format_number(result['arm_effect_at_centre']['p_value'])}, {result['arm_effect_at_centre']['sig']}).",
            f"beta_2 (Human-arm dose-response slope) = {format_number(result['dose_response_slope_human']['estimate'])} (p={format_number(result['dose_response_slope_human']['p_value'])}, {result['dose_response_slope_human']['sig']}).",
            f"beta_3 (interaction) = {format_number(result['interaction_term']['estimate'])} (p={format_number(result['interaction_term']['p_value'])}, {result['interaction_term']['sig']}).",
            result["interpretation"],
        ],
    )
    append_module(
        "Panel B - Nearest-Neighbour Matching",
        m2_results,
        lambda result: [
            f"Matched pairs = {result['n_matched_pairs']} within overlap [{format_number(result['overlap_lo'], 2)}, {format_number(result['overlap_hi'], 2)}].",
            f"Covariate-adjusted arm effect = {format_number(result['covariate_adjusted']['estimate'])} (SE {format_number(result['covariate_adjusted']['se'])}, p={format_number(result['covariate_adjusted']['p_value'])}, {result['covariate_adjusted']['sig']}).",
            f"Count+covariate-adjusted arm effect = {format_number(result['count_and_covariate_adjusted']['estimate'])} (p={format_number(result['count_and_covariate_adjusted']['p_value'])}, {result['count_and_covariate_adjusted']['sig']}).",
            result["interpretation"],
        ],
    )

    lines.extend(["", "## Panel C - Message-Level Content Audit", ""])
    lines.append(
        f"Message-level coded segments: Human n={quality_comparison.get('human_n_messages', 0)}, EPS-human n={quality_comparison.get('eps_n_messages', 0)}."
    )
    lines.extend(["", "| Feature | Human | EPS-human | EPS-Human diff | 95% CI | p |", "|---------|-------|-----------|----------------|--------|---|"])
    for item in quality_comparison.get("features", []):
        lines.append(
            f"| {item['label']} | {format_quality_value(item, item.get('human_value'))} | "
            f"{format_quality_value(item, item.get('eps_value'))} | "
            f"{format_quality_value(item, item.get('difference_eps_minus_human'))} | "
            f"[{format_quality_value(item, item.get('ci_low'))}, {format_quality_value(item, item.get('ci_high'))}] | "
            f"{format_number(item.get('p_value'))} {item.get('sig')} |"
        )
    lines.extend(["", quality_comparison.get("interpretation", ""), ""])

    eps_latency = arm_latency_stats.get("eps", {})
    human_latency = arm_latency_stats.get("human", {})
    mw = arm_latency_stats.get("mann_whitney", {})
    lines.extend(["## Panel D - Response Latency", ""])
    lines.extend(["| Arm | n | Median latency [IQR], min | Mean latency, min |", "|-----|---|---------------------------|-------------------|"])
    for item in (eps_latency, human_latency):
        lines.append(
            f"| {item.get('arm', 'NA')} | {item.get('n', 'NA')} | "
            f"{format_number(item.get('median'), 1)} [{format_number(item.get('p25'), 1)}, {format_number(item.get('p75'), 1)}] | "
            f"{format_number(item.get('mean'), 1)} |"
        )
    lines.append(
        f"\nMann-Whitney U={format_number(mw.get('u'), 1)}, z={format_number(mw.get('z'))}, p={format_number(mw.get('p'))} {pstar(mw.get('p')) if mw.get('p') is not None else 'NA'}."
    )

    lines.extend(["## Descriptive Statistics", "", "| Feature | Human mean (SD) | EPS mean (SD) |", "|---------|-----------------|---------------|"])
    for feature, values in desc.items():
        human = values["human"]
        eps = values["eps"]
        human_text = f"{human['mean']:.3f} ({human['sd']:.3f})" if human["mean"] is not None and human["sd"] is not None else "NA"
        eps_text = f"{eps['mean']:.3f} ({eps['sd']:.3f})" if eps["mean"] is not None and eps["sd"] is not None else "NA"
        lines.append(f"| {feature} | {human_text} | {eps_text} |")

    lines.extend(["", "## Plot Status", "", f"- {plot_status}", ""])
    return "\n".join(lines)


def derive_output_paths(outdir: Path, cohort: str) -> dict[str, Path]:
    prefix = f"{cohort}_frequency_control_content_audit"
    return {
        "summary_json": outdir / f"{prefix}_summary.json",
        "report_md": outdir / f"{prefix}_report.md",
        "results_xlsx": outdir / f"{prefix}_results.xlsx",
        "plot_png": outdir / f"{prefix}_plot.png",
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Run the checkin frequency-control/content-audit workflow for the updated "
            "Supplementary Table 3 logic."
        )
    )
    parser.add_argument("--cohort", choices=sorted(COHORT_CONFIGS), required=True, help="Clinical-trial cohort.")
    parser.add_argument("--human-file", type=Path, required=True, help="Human-arm workbook augmented with tagged-message counts.")
    parser.add_argument("--eps-file", type=Path, required=True, help="EPS-human-arm workbook augmented with tagged-message counts.")
    parser.add_argument("--human-report", type=Path, required=True, help="Human-arm JSON linkage report from build_checkin_dataset.py.")
    parser.add_argument("--eps-report", type=Path, required=True, help="EPS-human-arm JSON linkage report from build_checkin_dataset.py.")
    parser.add_argument("--human-chat-file", type=Path, required=True, help="Human-arm chat-export workbook.")
    parser.add_argument("--eps-chat-file", type=Path, required=True, help="EPS-human-arm chat-export workbook.")
    parser.add_argument("--outdir", type=Path, required=True, help="Directory for JSON/Markdown/Excel/PNG outputs.")
    parser.add_argument("--human-count-column", default=None, help="Optional override for the human-arm feedback count column.")
    parser.add_argument("--eps-count-column", default=None, help="Optional override for the EPS-human-arm feedback count column.")
    parser.add_argument("--human-keyword", default=None, help="Optional override for the human-arm tagged feedback keyword.")
    parser.add_argument("--eps-keyword", default=None, help="Optional override for the EPS-human-arm tagged feedback keyword.")
    parser.add_argument("--n-boot", type=int, default=DEFAULT_N_BOOT, help="Deprecated compatibility argument; bootstrap mediation is not run.")
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED, help="Random seed.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    config = COHORT_CONFIGS[args.cohort]
    human_count_candidates = [args.human_count_column] if args.human_count_column else HUMAN_COUNT_COLUMNS
    eps_count_candidates = [args.eps_count_column] if args.eps_count_column else EPS_COUNT_COLUMNS
    human_keyword = args.human_keyword or config.human_keyword
    eps_keyword = args.eps_keyword or config.eps_keyword
    output_paths = derive_output_paths(args.outdir, args.cohort)

    log("Loading participant datasets...")
    human_rows = load_arm_dataset(
        path=args.human_file,
        cohort=args.cohort,
        group_value=HUMAN_GROUP,
        count_candidates=human_count_candidates,
        arm_label="Human",
    )
    eps_rows = load_arm_dataset(
        path=args.eps_file,
        cohort=args.cohort,
        group_value=EPS_GROUP,
        count_candidates=eps_count_candidates,
        arm_label="EPS-human",
    )

    log("Loading linkage reports...")
    human_report = load_json(args.human_report)
    eps_report = load_json(args.eps_report)
    diagnostics = build_diagnostics(human_rows, eps_rows, human_report, eps_report)

    log("Extracting participant-level chat features...")
    human_features = parse_chat_features(chat_path=args.human_chat_file, participants=[row["user"] for row in human_rows], keyword=human_keyword, arm_label="Human")
    eps_features = parse_chat_features(chat_path=args.eps_chat_file, participants=[row["user"] for row in eps_rows], keyword=eps_keyword, arm_label="EPS-human")
    for row in human_rows:
        row.update(human_features.get(row["user"], {}))
    for row in eps_rows:
        row.update(eps_features.get(row["user"], {}))

    all_rows = human_rows + eps_rows

    log("Running message-level content audit...")
    human_quality_messages = extract_arm_quality_messages(chat_path=args.human_chat_file, keyword=human_keyword, arm_label="Human")
    eps_quality_messages = extract_arm_quality_messages(chat_path=args.eps_chat_file, keyword=eps_keyword, arm_label="EPS-human")
    quality_comparison = run_keyword_quality_comparison(human_quality_messages, eps_quality_messages)

    log("Extracting arm-level latency distributions from all keyword messages...")
    human_all_latencies = extract_arm_latencies_all_messages(chat_path=args.human_chat_file, keyword=human_keyword, arm_label="Human")
    eps_all_latencies = extract_arm_latencies_all_messages(chat_path=args.eps_chat_file, keyword=eps_keyword, arm_label="EPS-human")
    arm_latency_stats = {
        "human": arm_latency_descriptives(human_all_latencies, "Human"),
        "eps": arm_latency_descriptives(eps_all_latencies, "EPS-human"),
        "mann_whitney": mann_whitney_u(eps_all_latencies, human_all_latencies),
    }

    warnings = build_quality_warnings(diagnostics, all_rows)

    log("Running Panel A - Interaction...")
    m1_results = [run_interaction_model(all_rows, spec) for spec in config.outcome_specs]

    log("Running Panel B - Matching...")
    m2_results = [run_matching_analysis(all_rows, spec) for spec in config.outcome_specs]

    log("Attempting plot generation...")
    plot_status = maybe_build_plot(
        rows=all_rows,
        config=config,
        quality_comparison=quality_comparison,
        human_latencies=human_all_latencies,
        eps_latencies=eps_all_latencies,
        plot_path=output_paths["plot_png"],
    )

    desc = describe_features(all_rows)

    log("Writing Excel workbook...")
    write_results_workbook(
        config=config,
        diagnostics=diagnostics,
        warnings=warnings,
        rows=all_rows,
        desc=desc,
        m1_results=m1_results,
        m2_results=m2_results,
        quality_comparison=quality_comparison,
        arm_latency_stats=arm_latency_stats,
        human_keyword=human_keyword,
        eps_keyword=eps_keyword,
        n_boot=args.n_boot,
        seed=args.seed,
        plot_status=plot_status,
        output_xlsx=output_paths["results_xlsx"],
    )

    log("Writing Markdown report...")
    output_paths["report_md"].write_text(
        build_markdown_report(
            config=config,
            diagnostics=diagnostics,
            warnings=warnings,
            desc=desc,
            m1_results=m1_results,
            m2_results=m2_results,
            quality_comparison=quality_comparison,
            arm_latency_stats=arm_latency_stats,
            plot_status=plot_status,
        ),
        encoding="utf-8",
    )

    log("Writing JSON summary...")
    summary = {
        "analysis_scope": {
            "label": "frequency-control and content-audit analysis",
            "cohort": args.cohort,
            "description": "Updated article workflow using feedback-count interaction models, nearest-neighbour matching, message-level content audit, and response-latency descriptives. The content audit is construct-validity evidence, not formal causal mediation.",
            "deprecated_bootstrap_argument": args.n_boot,
            "random_seed": args.seed,
            "keywords": {"human": human_keyword, "eps": eps_keyword},
            "inputs": {
                "human_file": str(args.human_file),
                "eps_file": str(args.eps_file),
                "human_report_file": str(args.human_report),
                "eps_report_file": str(args.eps_report),
                "human_chat_file": str(args.human_chat_file),
                "eps_chat_file": str(args.eps_chat_file),
            },
        },
        "diagnostics": diagnostics,
        "warnings": warnings,
        "descriptive_statistics": desc,
        "panel_a_interaction": m1_results,
        "panel_b_matching": m2_results,
        "panel_c_content_audit": quality_comparison,
        "panel_d_latency": arm_latency_stats,
        "plot_status": plot_status,
        "outputs": {key: str(value) for key, value in output_paths.items()},
    }
    ensure_parent_dir(output_paths["summary_json"])
    output_paths["summary_json"].write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    log("Done.")


if __name__ == "__main__":
    main()
