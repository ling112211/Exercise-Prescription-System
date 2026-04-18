from __future__ import annotations

import argparse
import json
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


HEADER_FILL = PatternFill("solid", start_color="DBEAFE")
HEADER_FONT = Font(bold=True, color="1E3A8A")

WEIGHT_HUMAN_COUNTS = [1, 1, 2, 2, 3, 1, 2, 3]
WEIGHT_EPS_COUNTS = [2, 3, 3, 4, 4, 5, 3, 5]
GLYCEMIC_HUMAN_COUNTS = [1, 1, 2, 2, 2, 3, 1, 2]
GLYCEMIC_EPS_COUNTS = [2, 3, 3, 4, 2, 4, 3, 5]

WEIGHT_HUMAN_NAMES = [
    ("Ava Stone", "Ava"),
    ("Bella Reed", "Bella"),
    ("Clara Young", "Clara"),
    ("Diana Price", "Diana"),
    ("Elena Brooks", "Elena"),
    ("Fiona Hayes", "Fiona"),
    ("Grace West", "Grace"),
    ("Hazel Scott", "Hazel"),
]
WEIGHT_EPS_NAMES = [
    ("Ivy Turner", "Ivy"),
    ("Julia Hart", "Julia"),
    ("Kara Woods", "Kara"),
    ("Luna Ward", "Luna"),
    ("Mia Perry", "Mia"),
    ("Nora Lane", "Nora"),
    ("Olive Cole", "Olive"),
    ("Piper Bell", "Piper"),
]
GLYCEMIC_HUMAN_NAMES = [
    ("Quinn Baker", "Quinn"),
    ("Riley Long", "Riley"),
    ("Sofia Price", "Sofia"),
    ("Tessa Green", "Tessa"),
    ("Uma Fisher", "Uma"),
    ("Vera Knight", "Vera"),
    ("Wendy Ross", "Wendy"),
    ("Xena Moore", "Xena"),
]
GLYCEMIC_EPS_NAMES = [
    ("Yara Dixon", "Yara"),
    ("Zoe Kelly", "Zoe"),
    ("Aaron Fox", "Aaron"),
    ("Blake Reed", "Blake"),
    ("Carter Hall", "Carter"),
    ("Derek Wells", "Derek"),
    ("Ethan Shaw", "Ethan"),
    ("Felix Lowe", "Felix"),
]


def log(message: str) -> None:
    print(f"[synthetic-checkin] {message}", flush=True)


def ensure_parent_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def autosize(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for idx, width in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(width + 2, 32)


def write_table(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    ensure_parent_dir(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)
    wb.save(path)


def build_weight_rows(names: list[tuple[str, str]], counts: list[int], group: str, rng: np.random.Generator) -> tuple[list[list[object]], list[dict[str, object]]]:
    rows: list[list[object]] = []
    manifest: list[dict[str, object]] = []
    sexes = ["F", "F", "M", "F", "M", "F", "F", "M"]
    for index, ((full_name, short_name), feedback_count) in enumerate(zip(names, counts), start=1):
        age = int(rng.integers(32, 67))
        bmi = round(float(rng.uniform(23.0, 34.0)), 1)
        baseline_weight = round(float(rng.uniform(62.0, 96.0)), 1)
        group_bonus = 0.22 if group == "eps" else 0.0
        noise = float(rng.uniform(-0.08, 0.08))
        weight_loss_kg = round(0.45 + 0.18 * feedback_count + group_bonus + noise, 2)
        endline_weight = round(baseline_weight - weight_loss_kg, 2)
        weight_loss_pct = round(weight_loss_kg / baseline_weight, 4)
        participant_id = f"WL-{group.upper()}-{index:02d}"
        rows.append(
            [
                participant_id,
                full_name,
                age,
                sexes[index - 1],
                bmi,
                baseline_weight,
                endline_weight,
                weight_loss_kg,
                weight_loss_pct,
            ]
        )
        manifest.append(
            {
                "participant_id": participant_id,
                "user_nickname": full_name,
                "alias": short_name,
                "feedback_count": feedback_count,
            }
        )
    return rows, manifest


def build_glycemic_rows(names: list[tuple[str, str]], counts: list[int], group: str, rng: np.random.Generator) -> tuple[list[list[object]], list[dict[str, object]]]:
    rows: list[list[object]] = []
    manifest: list[dict[str, object]] = []
    sexes = ["F", "M", "F", "F", "M", "F", "M", "M"]
    for index, ((full_name, short_name), feedback_count) in enumerate(zip(names, counts), start=1):
        age = int(rng.integers(34, 70))
        bmi = round(float(rng.uniform(21.5, 33.5)), 1)
        baseline_fpg = round(float(rng.uniform(5.8, 9.6)), 2)
        group_bonus = 0.12 if group == "eps" else 0.0
        noise = float(rng.uniform(-0.05, 0.05))
        fpg_reduction = round(0.18 + 0.11 * feedback_count + group_bonus + noise, 2)
        endpoint_fpg = round(max(4.1, baseline_fpg - fpg_reduction), 2)
        participant_id = f"GLY-{group.upper()}-{index:02d}"
        rows.append(
            [
                participant_id,
                full_name,
                age,
                sexes[index - 1],
                bmi,
                baseline_fpg,
                endpoint_fpg,
            ]
        )
        manifest.append(
            {
                "participant_id": participant_id,
                "user_nickname": full_name,
                "alias": short_name,
                "feedback_count": feedback_count,
            }
        )
    return rows, manifest


def build_chat_rows(
    *,
    participants: list[dict[str, object]],
    cohort: str,
    keyword: str,
    arm_label: str,
    coach_name: str,
    start_day: int,
) -> list[list[object]]:
    rows: list[list[object]] = []
    minute_cursor = 0
    if cohort == "weight_loss":
        participant_templates = [
            "今天完成了{minutes}分钟快走和拉伸，晚餐控制得还可以。",
            "上午散步{minutes}分钟，晚上又补了力量训练，感觉状态不错。",
            "今天步数大概有{steps}步，还做了拉伸和深蹲。",
            "今天按照计划完成了有氧训练，运动后没有加餐。",
        ]
        feedback_templates = [
            "@{short_name} {keyword} 做得很好。根据你这周的记录，今天已经完成约{minutes}分钟活动，建议晚餐后再快走10分钟，继续保持步数。",
            "@{short_name} {keyword} 很棒，你最近执行得比较稳定。可以尝试把拉伸延长到8分钟，并注意控制晚餐主食比例。",
            "@{short_name} {keyword} 这次反馈重点看步数和训练连续性。建议明天把步行安排在饭后，保持节奏，你在进步。",
            "@{short_name} {keyword} 今天的完成度不错。结合你目前的状态，推荐继续快走加简单抗阻训练，注意补水和睡眠。",
        ]
        reply_templates = [
            "收到，谢谢老师，我明天继续打卡。",
            "好的，我会按建议把饭后步行也加上。",
            "明白了，我会继续保持今天的节奏。",
        ]
    else:
        participant_templates = [
            "今天餐后散步{minutes}分钟，空腹血糖感觉比前几天稳定一些。",
            "上午快走后记录了血糖，下午又做了轻度力量训练。",
            "今天控制了晚餐，总步数大概{steps}步，餐后活动也完成了。",
            "今天按照计划运动，餐后没有吃零食，准备继续观察空腹血糖。",
        ]
        feedback_templates = [
            "@{short_name} {keyword} 很好。根据你最近的血糖记录，建议餐后再步行10到15分钟，继续保持控糖节奏。",
            "@{short_name} {keyword} 做得不错，你这周的执行比较稳定。可以尝试把晚餐后活动固定下来，关注空腹和餐后血糖变化。",
            "@{short_name} {keyword} 今天反馈重点是控糖连续性。建议晚餐主食再略减一些，并保持饭后快走，你有明显进步。",
            "@{short_name} {keyword} 很棒，结合你目前的情况，推荐继续中等强度步行并记录空腹血糖，注意作息规律。",
        ]
        reply_templates = [
            "收到，我会继续按这个节奏控糖。",
            "好的，我明天把饭后步行时间再延长一点。",
            "明白了，我会继续记录空腹和餐后情况。",
        ]

    for index, participant in enumerate(participants, start=1):
        full_name = str(participant["user_nickname"])
        short_name = str(participant["alias"])
        feedback_count = int(participant["feedback_count"])
        participant_day = start_day + index
        minutes = 18 + index * 4
        steps = 4200 + index * 650

        rows.append(
            [
                f"2026-03-{participant_day:02d} 08:{minute_cursor:02d}",
                full_name,
                short_name,
                participant_templates[(index - 1) % len(participant_templates)].format(minutes=minutes, steps=steps),
            ]
        )
        minute_cursor = (minute_cursor + 7) % 60

        for turn in range(feedback_count):
            feedback_hour = min(21, 11 + turn * 3 + (index % 3))
            rows.append(
                [
                    f"2026-03-{participant_day:02d} {feedback_hour:02d}:{minute_cursor:02d}",
                    coach_name,
                    arm_label,
                    feedback_templates[(index + turn) % len(feedback_templates)].format(
                        short_name=short_name,
                        keyword=keyword,
                        minutes=minutes + turn * 5,
                        steps=steps + turn * 400,
                    ),
                ]
            )
            minute_cursor = (minute_cursor + 3) % 60

        rows.append(
            [
                f"2026-03-{participant_day:02d} 21:{minute_cursor:02d}",
                full_name,
                short_name,
                reply_templates[(index - 1) % len(reply_templates)],
            ]
        )
        minute_cursor = (minute_cursor + 5) % 60

    rows.append(
        [
            f"2026-03-{start_day + len(participants) + 2:02d} 20:15",
            coach_name,
            arm_label,
            f"@UnknownUser {keyword} 这是一条故意保留的未匹配示例。",
        ]
    )
    return rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate synthetic participant workbooks and chat-export workbooks for the checkin-analysis workflow."
    )
    parser.add_argument("--out-root", type=Path, default=Path("data/example/checkin"), help="Root directory for synthetic example files.")
    parser.add_argument("--seed", type=int, default=20260325, help="Random seed.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rng = np.random.default_rng(args.seed)

    manifest: dict[str, object] = {"seed": args.seed, "datasets": {}}

    weight_root = args.out_root / "weight_loss"
    gly_root = args.out_root / "glycemic"

    weight_headers = [
        "participant_id",
        "user_nickname",
        "age",
        "sex",
        "BMI",
        "baseline_weight_kg",
        "endline_weight_kg",
        "weight_loss_kg",
        "weight_loss_pct",
    ]
    gly_headers = [
        "participant_id",
        "user_nickname",
        "age",
        "sex",
        "BMI",
        "baseline_fpg_mmol",
        "endpoint_fpg_mmol",
    ]
    chat_headers = ["timestamp", "sender_name", "sender_group_nickname", "message_content"]

    weight_human_rows, weight_human_manifest = build_weight_rows(WEIGHT_HUMAN_NAMES, WEIGHT_HUMAN_COUNTS, "human", rng)
    weight_eps_rows, weight_eps_manifest = build_weight_rows(WEIGHT_EPS_NAMES, WEIGHT_EPS_COUNTS, "eps", rng)
    gly_human_rows, gly_human_manifest = build_glycemic_rows(GLYCEMIC_HUMAN_NAMES, GLYCEMIC_HUMAN_COUNTS, "human", rng)
    gly_eps_rows, gly_eps_manifest = build_glycemic_rows(GLYCEMIC_EPS_NAMES, GLYCEMIC_EPS_COUNTS, "eps", rng)

    write_table(weight_root / "human_arm.xlsx", weight_headers, weight_human_rows)
    write_table(weight_root / "eps_arm.xlsx", weight_headers, weight_eps_rows)
    write_table(gly_root / "human_arm.xlsx", gly_headers, gly_human_rows)
    write_table(gly_root / "eps_arm.xlsx", gly_headers, gly_eps_rows)

    write_table(
        weight_root / "human_chat_history.xlsx",
        chat_headers,
        build_chat_rows(
            participants=weight_human_manifest,
            cohort="weight_loss",
            keyword="#exercise feedback",
            arm_label="HumanCoach",
            coach_name="Coach Li",
            start_day=1,
        ),
    )
    write_table(
        weight_root / "eps_chat_history.xlsx",
        chat_headers,
        build_chat_rows(
            participants=weight_eps_manifest,
            cohort="weight_loss",
            keyword="#exercise feedback",
            arm_label="EPSCoach",
            coach_name="Coach Sun",
            start_day=10,
        ),
    )
    write_table(
        gly_root / "human_chat_history.xlsx",
        chat_headers,
        build_chat_rows(
            participants=gly_human_manifest,
            cohort="glycemic",
            keyword="#exercise feedback",
            arm_label="HumanCoach",
            coach_name="Coach Wu",
            start_day=1,
        ),
    )
    write_table(
        gly_root / "eps_chat_history.xlsx",
        chat_headers,
        build_chat_rows(
            participants=gly_eps_manifest,
            cohort="glycemic",
            keyword="#exercise feedback",
            arm_label="EPSCoach",
            coach_name="Coach Zhao",
            start_day=10,
        ),
    )

    manifest["datasets"] = {
        "weight_loss": {
            "human_participants": weight_human_manifest,
            "eps_participants": weight_eps_manifest,
            "files": {
                "human_arm": str(weight_root / "human_arm.xlsx"),
                "eps_arm": str(weight_root / "eps_arm.xlsx"),
                "human_chat_history": str(weight_root / "human_chat_history.xlsx"),
                "eps_chat_history": str(weight_root / "eps_chat_history.xlsx"),
            },
        },
        "glycemic": {
            "human_participants": gly_human_manifest,
            "eps_participants": gly_eps_manifest,
            "files": {
                "human_arm": str(gly_root / "human_arm.xlsx"),
                "eps_arm": str(gly_root / "eps_arm.xlsx"),
                "human_chat_history": str(gly_root / "human_chat_history.xlsx"),
                "eps_chat_history": str(gly_root / "eps_chat_history.xlsx"),
            },
        },
    }

    manifest_path = args.out_root / "synthetic_manifest.json"
    ensure_parent_dir(manifest_path)
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"Wrote synthetic data under {args.out_root}")


if __name__ == "__main__":
    main()
