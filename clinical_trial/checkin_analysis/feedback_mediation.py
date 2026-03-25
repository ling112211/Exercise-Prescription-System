from __future__ import annotations

import argparse
import contextlib
import io
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from statistics import NormalDist
from typing import Any

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HUMAN_GROUP = 0.0
EPS_GROUP = 1.0
STANDARD_NORMAL = NormalDist()

USER_COLUMNS = ["user_nickname", "participant_nickname", "用户昵称"]
AGE_COLUMNS = ["age", "年龄"]
SEX_COLUMNS = ["sex", "gender", "性别"]
BMI_COLUMNS = ["BMI", "bmi", "体重指数"]
HUMAN_COUNT_COLUMNS = ["daily_activity_checkin_count", "包含#日常活动打卡消息总数"]
EPS_COUNT_COLUMNS = ["exercise_feedback_count", "包含#运动点评消息总数"]

WEIGHT_BASELINE_COLUMNS = ["baseline_weight_kg", "baseline_weight", "weight_baseline", "入营体重", "入营体重kg", "入营体重（档案）"]
WEIGHT_ENDLINE_COLUMNS = ["endline_weight_kg", "endline_weight", "weight_endline", "出营体重", "出营体重kg"]
WEIGHT_LOSS_KG_COLUMNS = ["weight_loss_kg", "减重数"]
WEIGHT_LOSS_PCT_COLUMNS = ["weight_loss_pct", "减重比率"]

GLYCEMIC_BASELINE_COLUMNS = ["baseline_fpg_mmol", "baseline_fpg", "FPG0", "入营空腹", "Baseline fasting glucose", "Fasting glucose"]
GLYCEMIC_ENDLINE_COLUMNS = ["endpoint_fpg_mmol", "endline_fpg_mmol", "endpoint_fpg", "结营空腹"]

LINE_FILL = PatternFill("solid", start_color="1D4ED8")
SECTION_FILL = PatternFill("solid", start_color="EFF6FF")
THIN = Side(style="thin", color="CBD5E1")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, color="1D4ED8")
LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
SPACE_TRANSLATION = str.maketrans({"\u00a0": " ", "\u200b": " ", "\ufeff": " "})
QUOTE_TRANSLATION = str.maketrans({"“": '"', "”": '"', "‘": "'", "’": "'"})


@dataclass(frozen=True)
class OutcomeSpec:
    key: str
    title: str
    description: str
    note: str


WEIGHT_OUTCOME_SPECS = [
    OutcomeSpec(
        key="endline_weight",
        title="Primary ANCOVA-Style Endpoint Model",
        description="Outcome = day-21 body weight (kg); lower values are better.",
        note="Negative treatment/direct effects favor EPS because they imply lower endline body weight after baseline adjustment.",
    ),
    OutcomeSpec(
        key="weight_loss_kg",
        title="Weight-Loss Sensitivity Model",
        description="Outcome = baseline body weight minus day-21 body weight (kg); higher values are better.",
        note="Positive treatment/direct effects favor EPS because they imply larger absolute weight loss.",
    ),
    OutcomeSpec(
        key="weight_loss_pct",
        title="Percent Weight-Loss Sensitivity Model",
        description="Outcome = percent weight loss from baseline to day-21; higher values are better.",
        note="Positive treatment/direct effects favor EPS because they imply larger percent weight loss.",
    ),
]

GLYCEMIC_OUTCOME_SPECS = [
    OutcomeSpec(
        key="endpoint_fpg",
        title="Primary ANCOVA-Style Endpoint Model",
        description="Outcome = day-21 fasting glucose (mmol/L); lower values are better.",
        note="Negative treatment/direct effects favor EPS because they imply lower endline fasting glucose after baseline adjustment.",
    ),
    OutcomeSpec(
        key="fpg_reduction",
        title="Change-Score Sensitivity Model",
        description="Outcome = baseline fasting glucose minus day-21 fasting glucose (mmol/L); higher values are better.",
        note="Positive treatment/direct effects favor EPS because they imply larger fasting-glucose reduction.",
    ),
]


def log(message: str) -> None:
    print(f"[checkin-mediation] {message}", flush=True)


def ensure_parent_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def normalize_text(text: object) -> str:
    text = str(text).translate(SPACE_TRANSLATION).translate(QUOTE_TRANSLATION)
    return re.sub(r"[ \t\r\f\v]+", " ", text).strip()


def clean_name(name: object | None) -> str:
    if name is None:
        return ""
    return normalize_text(name).strip(" @\"'`:,;:：")


def pick_header_name(
    header: list[str],
    candidates: list[str],
    *,
    required: bool,
    label: str,
    path: Path,
) -> str | None:
    exact = {str(value).strip(): str(value).strip() for value in header if value is not None}
    lower = {key.lower(): key for key in exact}
    for candidate in candidates:
        candidate_clean = str(candidate).strip()
        if candidate_clean in exact:
            return exact[candidate_clean]
        if candidate_clean.lower() in lower:
            return lower[candidate_clean.lower()]
    if required:
        raise KeyError(f"Could not find {label} column in {path.name}. Candidates={candidates}")
    return None


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("%", "")
    try:
        return float(text)
    except ValueError:
        return None


def to_clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def sex_to_female_indicator(value: Any) -> float | None:
    text = to_clean_text(value).strip().lower()
    if not text:
        return None
    female_tokens = {"f", "female", "woman", "女", "0", "2"}
    male_tokens = {"m", "male", "man", "男", "1"}
    if text in female_tokens:
        return 1.0
    if text in male_tokens:
        return 0.0
    return None


def format_number(value: float | int | None, digits: int = 4) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "NA"
    if isinstance(value, int):
        return str(value)
    return f"{value:.{digits}f}"


def pstar(p_value: float | None) -> str:
    if p_value is None or math.isnan(p_value):
        return "NA"
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    return "ns"


def normal_two_sided_p_from_z(z_value: float | None) -> float | None:
    if z_value is None or math.isnan(z_value):
        return None
    return 2.0 * (1.0 - STANDARD_NORMAL.cdf(abs(z_value)))


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def load_arm_dataset(
    *,
    path: Path,
    cohort: str,
    group_value: float,
    count_candidates: list[str],
    arm_label: str,
) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header = [str(value) if value is not None else "" for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]

    user_column = pick_header_name(header, USER_COLUMNS, required=True, label="participant-name", path=path)
    age_column = pick_header_name(header, AGE_COLUMNS, required=True, label="age", path=path)
    sex_column = pick_header_name(header, SEX_COLUMNS, required=True, label="sex", path=path)
    bmi_column = pick_header_name(header, BMI_COLUMNS, required=True, label="BMI", path=path)
    count_column = pick_header_name(header, count_candidates, required=True, label="feedback-count", path=path)

    user_idx = header.index(user_column)
    age_idx = header.index(age_column)
    sex_idx = header.index(sex_column)
    bmi_idx = header.index(bmi_column)
    count_idx = header.index(count_column)

    baseline_idx = None
    endline_idx = None
    aux_kg_idx = None
    aux_pct_idx = None
    if cohort == "weight_loss":
        baseline_column = pick_header_name(header, WEIGHT_BASELINE_COLUMNS, required=True, label="baseline-weight", path=path)
        baseline_idx = header.index(baseline_column)
        endline_column = pick_header_name(header, WEIGHT_ENDLINE_COLUMNS, required=False, label="endline-weight", path=path)
        endline_idx = header.index(endline_column) if endline_column else None
        aux_kg_column = pick_header_name(header, WEIGHT_LOSS_KG_COLUMNS, required=False, label="weight-loss-kg", path=path)
        aux_pct_column = pick_header_name(header, WEIGHT_LOSS_PCT_COLUMNS, required=False, label="weight-loss-pct", path=path)
        aux_kg_idx = header.index(aux_kg_column) if aux_kg_column else None
        aux_pct_idx = header.index(aux_pct_column) if aux_pct_column else None
        if endline_idx is None and aux_kg_idx is None:
            raise KeyError(
                f"{path.name} must include at least one of {WEIGHT_ENDLINE_COLUMNS} or {WEIGHT_LOSS_KG_COLUMNS}."
            )
    else:
        baseline_column = pick_header_name(header, GLYCEMIC_BASELINE_COLUMNS, required=True, label="baseline-fpg", path=path)
        endline_column = pick_header_name(header, GLYCEMIC_ENDLINE_COLUMNS, required=True, label="endline-fpg", path=path)
        baseline_idx = header.index(baseline_column)
        endline_idx = header.index(endline_column)

    dataset: list[dict[str, Any]] = []
    seen_users: set[str] = set()
    duplicate_rows_skipped = 0
    blank_streak = 0
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in row):
            blank_streak += 1
            if blank_streak >= 200:
                break
            continue

        raw_user = to_clean_text(row[user_idx])
        user = clean_name(raw_user)
        if not user:
            blank_streak += 1
            if blank_streak >= 200:
                break
            continue

        blank_streak = 0
        if user in seen_users:
            duplicate_rows_skipped += 1
            continue
        seen_users.add(user)

        age = to_float(row[age_idx])
        bmi = to_float(row[bmi_idx])
        sex_f = sex_to_female_indicator(row[sex_idx])
        feedback_count = to_float(row[count_idx]) or 0.0

        record = {
            "arm": arm_label,
            "group": group_value,
            "user": user,
            "user_raw": raw_user,
            "feedback_count": feedback_count,
            "received_any_feedback": 1.0 if feedback_count >= 1 else 0.0,
            "age": age,
            "bmi": bmi,
            "sex_f": sex_f,
        }

        if cohort == "weight_loss":
            baseline_weight = to_float(row[baseline_idx])
            endline_weight = to_float(row[endline_idx]) if endline_idx is not None else None
            weight_loss_kg = to_float(row[aux_kg_idx]) if aux_kg_idx is not None else None
            weight_loss_pct = to_float(row[aux_pct_idx]) if aux_pct_idx is not None else None

            if weight_loss_pct is not None and abs(weight_loss_pct) > 1.0:
                weight_loss_pct = weight_loss_pct / 100.0
            if weight_loss_kg is None and baseline_weight is not None and endline_weight is not None:
                weight_loss_kg = baseline_weight - endline_weight
            if endline_weight is None and baseline_weight is not None and weight_loss_kg is not None:
                endline_weight = baseline_weight - weight_loss_kg
            if weight_loss_pct is None and baseline_weight not in (None, 0, 0.0) and weight_loss_kg is not None:
                weight_loss_pct = weight_loss_kg / baseline_weight

            record.update(
                {
                    "baseline_weight": baseline_weight,
                    "endline_weight": endline_weight,
                    "weight_loss_kg": weight_loss_kg,
                    "weight_loss_pct": weight_loss_pct,
                }
            )
        else:
            baseline_fpg = to_float(row[baseline_idx])
            endpoint_fpg = to_float(row[endline_idx])
            record.update(
                {
                    "baseline_fpg": baseline_fpg,
                    "endpoint_fpg": endpoint_fpg,
                    "fpg_reduction": None if baseline_fpg is None or endpoint_fpg is None else baseline_fpg - endpoint_fpg,
                }
            )

        dataset.append(record)

    workbook.close()
    if duplicate_rows_skipped:
        log(f"{arm_label}: skipped {duplicate_rows_skipped} duplicate row(s) after nickname normalization.")
    return dataset


def build_diagnostics(
    human_rows: list[dict[str, Any]],
    eps_rows: list[dict[str, Any]],
    human_report: dict[str, Any],
    eps_report: dict[str, Any],
) -> list[dict[str, Any]]:
    diagnostics: list[dict[str, Any]] = []
    for arm_label, rows, report in (
        ("Human", human_rows, human_report),
        ("EPS-human", eps_rows, eps_report),
    ):
        feedback_total = int(round(sum(row["feedback_count"] for row in rows)))
        tagged_total = int(report.get("keyword_message_count", 0))
        reported_feedback_total = int(report.get("resolved_message_count", 0))
        unresolved_total = int(report.get("unresolved_message_count", 0))
        diagnostics.append(
            {
                "arm": arm_label,
                "participant_count": len(rows),
                "feedback_total_from_output_excel": feedback_total,
                "feedback_total_from_report": reported_feedback_total,
                "tagged_message_total_from_report": tagged_total,
                "unresolved_message_count_from_report": unresolved_total,
                "output_matches_report_feedback_total": feedback_total == reported_feedback_total,
            }
        )
    return diagnostics


def build_quality_warnings(diagnostics: list[dict[str, Any]]) -> list[str]:
    warnings: list[str] = []
    for item in diagnostics:
        arm = item["arm"]
        if not item["output_matches_report_feedback_total"]:
            warnings.append(f"{arm}: feedback totals in the output workbook do not match the JSON linkage report.")
    return warnings


def rows_to_matrix(rows: list[dict[str, Any]], cohort: str, outcome_key: str) -> dict[str, np.ndarray]:
    baseline_key = "baseline_weight" if cohort == "weight_loss" else "baseline_fpg"
    filtered = [
        row
        for row in rows
        if row.get(outcome_key) is not None
        and row.get("feedback_count") is not None
        and row.get(baseline_key) is not None
        and row.get("age") is not None
        and row.get("bmi") is not None
        and row.get("sex_f") is not None
    ]
    if len(filtered) < 8:
        raise ValueError(f"Too few complete cases for outcome {outcome_key!r}: {len(filtered)}")

    return {
        "rows": filtered,
        "X": np.array([row["group"] for row in filtered], dtype=float),
        "M": np.array([row["feedback_count"] for row in filtered], dtype=float),
        "Mbin": np.array([row["received_any_feedback"] for row in filtered], dtype=float),
        "Y": np.array([row[outcome_key] for row in filtered], dtype=float),
        "covars": np.array(
            [[row[baseline_key], row["age"], row["bmi"], row["sex_f"]] for row in filtered],
            dtype=float,
        ),
        "groups": np.array([row["group"] for row in filtered], dtype=float),
    }


def with_intercept(*columns: np.ndarray) -> np.ndarray:
    n_rows = len(columns[0])
    return np.column_stack([np.ones(n_rows)] + [np.asarray(column, dtype=float) for column in columns])


def fit_ols(y: np.ndarray, design: np.ndarray) -> dict[str, Any]:
    y = np.asarray(y, dtype=float)
    design = np.asarray(design, dtype=float)
    n_obs, n_params = design.shape
    with np.errstate(divide="ignore", invalid="ignore", over="ignore"):
        xtx_inv = np.linalg.pinv(design.T @ design)
        beta = xtx_inv @ design.T @ y
        fitted = design @ beta
        resid = y - fitted

        hat_diag = np.sum(design * (design @ xtx_inv), axis=1)
        hat_diag = np.clip(hat_diag, 0.0, 1.0 - 1e-8)
        hat_denom = np.clip(1.0 - hat_diag, 1e-8, None)
        scale = np.square(resid / hat_denom)
        meat = design.T @ (design * scale[:, None])
        cov_hc3 = xtx_inv @ meat @ xtx_inv
    se = np.sqrt(np.maximum(np.diag(cov_hc3), 0.0))

    df = max(n_obs - n_params, 1)
    t_stat = np.divide(beta, se, out=np.full_like(beta, np.nan), where=se > 0)
    p_value = np.array([normal_two_sided_p_from_z(float(value)) for value in t_stat], dtype=float)

    rss = float(np.sum(resid**2))
    tss = float(np.sum((y - np.mean(y)) ** 2))
    r2 = float("nan") if tss <= 0 else 1.0 - rss / tss
    return {
        "beta": beta,
        "se": se,
        "p": p_value,
        "resid": resid,
        "fitted": fitted,
        "r2": r2,
        "df_resid": df,
        "n_obs": n_obs,
    }


def fit_ols_plain(y: np.ndarray, design: np.ndarray) -> np.ndarray:
    beta, *_ = np.linalg.lstsq(design, y, rcond=None)
    return beta


def bootstrap_indirect_effect(
    *,
    x: np.ndarray,
    m: np.ndarray,
    y: np.ndarray,
    covars: np.ndarray,
    n_boot: int,
    seed: int,
) -> dict[str, Any]:
    rng = np.random.default_rng(seed)
    n_obs = len(y)
    draws = np.empty(n_boot, dtype=float)
    for index in range(n_boot):
        sample_idx = rng.integers(0, n_obs, size=n_obs)
        xb = x[sample_idx]
        mb = m[sample_idx]
        yb = y[sample_idx]
        cb = covars[sample_idx]
        a_beta = fit_ols_plain(mb, with_intercept(xb, cb))
        b_beta = fit_ols_plain(yb, with_intercept(xb, mb, cb))
        draws[index] = float(a_beta[1] * b_beta[2])

    ci_low, ci_high = np.percentile(draws, [2.5, 97.5])
    return {
        "draws": draws,
        "ci_low": float(ci_low),
        "ci_high": float(ci_high),
        "significant": not (ci_low <= 0.0 <= ci_high),
    }


def rank_average_ties(values: np.ndarray) -> np.ndarray:
    order = np.argsort(values, kind="mergesort")
    ranks = np.empty(len(values), dtype=float)
    sorted_values = values[order]
    start = 0
    while start < len(values):
        end = start + 1
        while end < len(values) and sorted_values[end] == sorted_values[start]:
            end += 1
        average_rank = (start + 1 + end) / 2.0
        ranks[order[start:end]] = average_rank
        start = end
    return ranks


def pearson_correlation(x: np.ndarray, y: np.ndarray) -> float | None:
    x_centered = x - np.mean(x)
    y_centered = y - np.mean(y)
    denom = math.sqrt(float(np.sum(x_centered**2) * np.sum(y_centered**2)))
    if denom <= 0:
        return None
    return float(np.sum(x_centered * y_centered) / denom)


def fit_spearman(x: np.ndarray, y: np.ndarray) -> dict[str, float | None]:
    if len(x) < 3 or np.all(x == x[0]) or np.all(y == y[0]):
        return {"rho": None, "p": None}
    x_rank = rank_average_ties(x)
    y_rank = rank_average_ties(y)
    rho = pearson_correlation(x_rank, y_rank)
    if rho is None:
        return {"rho": None, "p": None}
    if abs(rho) >= 1.0:
        return {"rho": float(rho), "p": 0.0}
    n_obs = len(x)
    t_like = rho * math.sqrt((n_obs - 2.0) / max(1e-12, 1.0 - rho**2))
    return {"rho": float(rho), "p": normal_two_sided_p_from_z(t_like)}


def fit_linear_slope(x: np.ndarray, y: np.ndarray) -> dict[str, float | None]:
    if len(x) < 3 or np.all(x == x[0]):
        return {"slope": None, "intercept": None}
    x_mean = float(np.mean(x))
    y_mean = float(np.mean(y))
    ss_x = float(np.sum((x - x_mean) ** 2))
    if ss_x <= 0:
        return {"slope": None, "intercept": None}
    slope = float(np.sum((x - x_mean) * (y - y_mean)) / ss_x)
    intercept = y_mean - slope * x_mean
    return {"slope": slope, "intercept": float(intercept)}


def run_mediation_spec(
    *,
    rows: list[dict[str, Any]],
    cohort: str,
    spec: OutcomeSpec,
    n_boot: int,
    seed: int,
) -> dict[str, Any]:
    arrays = rows_to_matrix(rows, cohort, spec.key)
    x = arrays["X"]
    m = arrays["M"]
    m_bin = arrays["Mbin"]
    y = arrays["Y"]
    covars = arrays["covars"]
    groups = arrays["groups"]

    total_model = fit_ols(y, with_intercept(x, covars))
    mediator_model = fit_ols(m, with_intercept(x, covars))
    direct_model = fit_ols(y, with_intercept(x, m, covars))

    c_total = float(total_model["beta"][1])
    a_path = float(mediator_model["beta"][1])
    c_prime = float(direct_model["beta"][1])
    b_path = float(direct_model["beta"][2])
    indirect = a_path * b_path
    proportion = indirect / c_total if abs(c_total) > 1e-12 else None

    boot = bootstrap_indirect_effect(x=x, m=m, y=y, covars=covars, n_boot=n_boot, seed=seed)

    mediator_binary_model = fit_ols(m_bin, with_intercept(x, covars))
    direct_binary_model = fit_ols(y, with_intercept(x, m_bin, covars))
    binary_indirect = float(mediator_binary_model["beta"][1] * direct_binary_model["beta"][2])
    boot_binary = bootstrap_indirect_effect(
        x=x,
        m=m_bin,
        y=y,
        covars=covars,
        n_boot=n_boot,
        seed=seed + 1,
    )

    human_mask = groups == HUMAN_GROUP
    eps_mask = groups == EPS_GROUP
    human_corr = fit_spearman(m[human_mask], y[human_mask])
    eps_corr = fit_spearman(m[eps_mask], y[eps_mask])
    human_slope = fit_linear_slope(m[human_mask], y[human_mask])
    eps_slope = fit_linear_slope(m[eps_mask], y[eps_mask])

    return {
        "outcome_key": spec.key,
        "outcome_title": spec.title,
        "outcome_description": spec.description,
        "direction_note": spec.note,
        "n_obs": int(len(y)),
        "human_n": int(np.sum(human_mask)),
        "eps_n": int(np.sum(eps_mask)),
        "total_effect": {
            "estimate": c_total,
            "se_hc3": float(total_model["se"][1]),
            "p_value": float(total_model["p"][1]),
            "r2": float(total_model["r2"]),
        },
        "a_path": {
            "estimate": a_path,
            "se_hc3": float(mediator_model["se"][1]),
            "p_value": float(mediator_model["p"][1]),
            "r2": float(mediator_model["r2"]),
        },
        "b_path": {
            "estimate": b_path,
            "se_hc3": float(direct_model["se"][2]),
            "p_value": float(direct_model["p"][2]),
        },
        "direct_effect": {
            "estimate": c_prime,
            "se_hc3": float(direct_model["se"][1]),
            "p_value": float(direct_model["p"][1]),
            "r2": float(direct_model["r2"]),
        },
        "indirect_effect": {
            "estimate": float(indirect),
            "ci_low": boot["ci_low"],
            "ci_high": boot["ci_high"],
            "significant": bool(boot["significant"]),
            "proportion_of_total": proportion,
        },
        "binary_mediator_sensitivity": {
            "definition": "received >=1 actual tagged feedback message",
            "a_path_estimate": float(mediator_binary_model["beta"][1]),
            "a_path_p_value": float(mediator_binary_model["p"][1]),
            "b_path_estimate": float(direct_binary_model["beta"][2]),
            "b_path_p_value": float(direct_binary_model["p"][2]),
            "indirect_effect": binary_indirect,
            "ci_low": boot_binary["ci_low"],
            "ci_high": boot_binary["ci_high"],
            "significant": bool(boot_binary["significant"]),
        },
        "within_arm_dose_response": {
            "human": {
                "spearman_rho": human_corr["rho"],
                "spearman_p_value": human_corr["p"],
                "linear_slope": human_slope["slope"],
            },
            "eps_human": {
                "spearman_rho": eps_corr["rho"],
                "spearman_p_value": eps_corr["p"],
                "linear_slope": eps_slope["slope"],
            },
        },
        "bootstrap_draws": boot["draws"].tolist(),
        "binary_bootstrap_draws": boot_binary["draws"].tolist(),
    }


def build_result_interpretation(result: dict[str, Any]) -> list[str]:
    indirect = result["indirect_effect"]
    direct = result["direct_effect"]
    total = result["total_effect"]
    return [
        f"{result['outcome_title']}: n={result['n_obs']} (Human={result['human_n']}, EPS-human={result['eps_n']}).",
        f"Total effect estimate = {format_number(total['estimate'])} (HC3 SE {format_number(total['se_hc3'])}, p={format_number(total['p_value'])}).",
        f"Path a estimate = {format_number(result['a_path']['estimate'])} (p={format_number(result['a_path']['p_value'])}); path b estimate = {format_number(result['b_path']['estimate'])} (p={format_number(result['b_path']['p_value'])}).",
        f"Indirect effect = {format_number(indirect['estimate'])}, bootstrap 95% CI [{format_number(indirect['ci_low'])}, {format_number(indirect['ci_high'])}], significant={indirect['significant']}.",
        f"Direct effect estimate = {format_number(direct['estimate'])} (p={format_number(direct['p_value'])}).",
        result["direction_note"],
    ]


def maybe_build_plot(
    *,
    dataset_rows: list[dict[str, Any]],
    cohort: str,
    primary_result: dict[str, Any],
    plot_path: Path,
) -> str:
    try:
        with contextlib.redirect_stderr(io.StringIO()), contextlib.redirect_stdout(io.StringIO()):
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
    except Exception as exc:
        return f"Skipped plot generation because matplotlib is unavailable: {exc}"

    primary_rows = rows_to_matrix(dataset_rows, cohort, primary_result["outcome_key"])
    x = primary_rows["M"]
    y = primary_rows["Y"]
    groups = primary_rows["groups"]
    human_mask = groups == HUMAN_GROUP
    eps_mask = groups == EPS_GROUP

    fig, axes = plt.subplots(1, 2, figsize=(12, 5.5))
    fig.patch.set_facecolor("white")
    colors = {"Human": "#2563EB", "EPS-human": "#EA580C"}

    for mask, label in ((human_mask, "Human"), (eps_mask, "EPS-human")):
        x_arm = x[mask]
        y_arm = y[mask]
        axes[0].scatter(
            x_arm,
            y_arm,
            label=label,
            color=colors[label],
            alpha=0.8,
            s=50,
            edgecolors="white",
            linewidth=0.8,
        )
        slope_info = fit_linear_slope(x_arm, y_arm)
        if slope_info["slope"] is not None and slope_info["intercept"] is not None:
            x_grid = np.linspace(float(np.min(x_arm)), float(np.max(x_arm)), 100)
            y_grid = slope_info["slope"] * x_grid + slope_info["intercept"]
            axes[0].plot(x_grid, y_grid, color=colors[label], linewidth=2)

    axes[0].set_xlabel("Actual feedback count from chat history")
    axes[0].set_ylabel(primary_result["outcome_description"])
    axes[0].set_title("Within-arm dose-response")
    axes[0].legend()
    axes[0].spines[["top", "right"]].set_visible(False)

    draws = np.array(primary_result["bootstrap_draws"], dtype=float)
    axes[1].hist(draws, bins=60, color="#2563EB", alpha=0.7, edgecolor="white")
    axes[1].axvline(primary_result["indirect_effect"]["estimate"], color="black", linewidth=2, label="Observed indirect")
    axes[1].axvline(primary_result["indirect_effect"]["ci_low"], color="red", linestyle="--", linewidth=1.5, label="95% CI")
    axes[1].axvline(primary_result["indirect_effect"]["ci_high"], color="red", linestyle="--", linewidth=1.5)
    axes[1].axvline(0.0, color="#6B7280", linestyle=":", linewidth=1.2)
    axes[1].set_xlabel("Bootstrap indirect effect (a*b)")
    axes[1].set_ylabel("Frequency")
    axes[1].set_title("Indirect-effect bootstrap distribution")
    axes[1].legend()
    axes[1].spines[["top", "right"]].set_visible(False)

    fig.suptitle("Exploratory mediation analysis")
    fig.tight_layout()
    ensure_parent_dir(plot_path)
    fig.savefig(plot_path, dpi=180, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return f"Saved plot to {plot_path.name}"


def write_cell(
    ws,
    row: int,
    col: int,
    value: Any,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    align: Alignment | None = None,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    cell.border = CELL_BORDER


def autosize_columns(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for column_idx, width in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=column_idx).column_letter].width = min(width + 2, 48)


def dataset_headers_for_cohort(cohort: str) -> list[str]:
    if cohort == "weight_loss":
        return [
            "arm",
            "group",
            "user",
            "feedback_count",
            "received_any_feedback",
            "baseline_weight",
            "endline_weight",
            "weight_loss_kg",
            "weight_loss_pct",
            "age",
            "bmi",
            "sex_f",
        ]
    return [
        "arm",
        "group",
        "user",
        "feedback_count",
        "received_any_feedback",
        "baseline_fpg",
        "endpoint_fpg",
        "fpg_reduction",
        "age",
        "bmi",
        "sex_f",
    ]


def write_results_workbook(
    *,
    cohort: str,
    diagnostics: list[dict[str, Any]],
    warnings: list[str],
    dataset_rows: list[dict[str, Any]],
    analysis_results: list[dict[str, Any]],
    plot_status: str,
    output_xlsx: Path,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    ws_diag = workbook.create_sheet("Diagnostics")
    diag_headers = [
        "arm",
        "participant_count",
        "feedback_total_from_output_excel",
        "feedback_total_from_report",
        "tagged_message_total_from_report",
        "unresolved_message_count_from_report",
        "output_matches_report_feedback_total",
    ]
    for col_idx, header in enumerate(diag_headers, start=1):
        write_cell(ws_diag, 1, col_idx, header, font=HEADER_FONT, fill=LINE_FILL, align=CENTER)
    for row_idx, item in enumerate(diagnostics, start=2):
        for col_idx, header in enumerate(diag_headers, start=1):
            write_cell(ws_diag, row_idx, col_idx, item.get(header), align=CENTER)
    next_row = len(diagnostics) + 3
    write_cell(ws_diag, next_row, 1, "warnings", font=SECTION_FONT, fill=SECTION_FILL, align=LEFT)
    for offset, message in enumerate(warnings or ["No warnings detected."], start=1):
        write_cell(ws_diag, next_row + offset, 1, message, align=LEFT)
    write_cell(ws_diag, next_row + len(warnings or ["No warnings detected."]) + 2, 1, plot_status, align=LEFT)
    autosize_columns(ws_diag)

    ws_data = workbook.create_sheet("Participant Dataset")
    data_headers = dataset_headers_for_cohort(cohort)
    for col_idx, header in enumerate(data_headers, start=1):
        write_cell(ws_data, 1, col_idx, header, font=HEADER_FONT, fill=LINE_FILL, align=CENTER)
    for row_idx, item in enumerate(dataset_rows, start=2):
        for col_idx, header in enumerate(data_headers, start=1):
            write_cell(ws_data, row_idx, col_idx, item.get(header), align=CENTER)
    autosize_columns(ws_data)

    ws_results = workbook.create_sheet("Mediation Results")
    result_headers = ["Outcome", "Parameter", "Estimate", "HC3_SE", "p_value", "Sig", "CI_low", "CI_high", "Note"]
    for col_idx, header in enumerate(result_headers, start=1):
        write_cell(ws_results, 1, col_idx, header, font=HEADER_FONT, fill=LINE_FILL, align=CENTER)

    row_cursor = 2
    for result in analysis_results:
        write_cell(ws_results, row_cursor, 1, result["outcome_title"], font=SECTION_FONT, fill=SECTION_FILL, align=LEFT)
        row_cursor += 1
        result_rows = [
            ("total effect c", result["total_effect"]["estimate"], result["total_effect"]["se_hc3"], result["total_effect"]["p_value"], "", "", "Outcome ~ treatment + covariates"),
            ("path a", result["a_path"]["estimate"], result["a_path"]["se_hc3"], result["a_path"]["p_value"], "", "", "Mediator ~ treatment + covariates"),
            ("path b", result["b_path"]["estimate"], result["b_path"]["se_hc3"], result["b_path"]["p_value"], "", "", "Outcome ~ treatment + mediator + covariates"),
            ("direct effect c'", result["direct_effect"]["estimate"], result["direct_effect"]["se_hc3"], result["direct_effect"]["p_value"], "", "", "Treatment effect controlling for mediator"),
            ("indirect effect a*b", result["indirect_effect"]["estimate"], None, None, result["indirect_effect"]["ci_low"], result["indirect_effect"]["ci_high"], "Bootstrap percentile CI"),
            ("binary mediator indirect", result["binary_mediator_sensitivity"]["indirect_effect"], None, None, result["binary_mediator_sensitivity"]["ci_low"], result["binary_mediator_sensitivity"]["ci_high"], "Binary mediator sensitivity"),
        ]
        for parameter, estimate, se, p_value, ci_low, ci_high, note in result_rows:
            write_cell(ws_results, row_cursor, 1, result["outcome_title"], align=LEFT)
            write_cell(ws_results, row_cursor, 2, parameter, align=LEFT)
            write_cell(ws_results, row_cursor, 3, estimate, align=CENTER)
            write_cell(ws_results, row_cursor, 4, se, align=CENTER)
            write_cell(ws_results, row_cursor, 5, p_value, align=CENTER)
            write_cell(ws_results, row_cursor, 6, pstar(p_value if isinstance(p_value, (int, float)) else None), align=CENTER)
            write_cell(ws_results, row_cursor, 7, ci_low, align=CENTER)
            write_cell(ws_results, row_cursor, 8, ci_high, align=CENTER)
            write_cell(ws_results, row_cursor, 9, note, align=LEFT)
            row_cursor += 1

        write_cell(ws_results, row_cursor, 2, "direction note", align=LEFT)
        write_cell(ws_results, row_cursor, 9, result["direction_note"], align=LEFT)
        row_cursor += 2

    autosize_columns(ws_results)
    ensure_parent_dir(output_xlsx)
    workbook.save(output_xlsx)


def build_markdown_report(
    *,
    cohort: str,
    diagnostics: list[dict[str, Any]],
    warnings: list[str],
    analysis_results: list[dict[str, Any]],
    plot_status: str,
) -> str:
    cohort_label = "weight-loss cohort" if cohort == "weight_loss" else "glycemic-control cohort"
    lines = [
        "# Exploratory Mediation Memo",
        "",
        "## Scope",
        "",
        f"This script uses actual feedback counts reconstructed from chat-export records in the {cohort_label} as the mediator.",
        "Feedback counts come from keyword matching plus participant linking in the chat history. The linkage report records any tagged messages that remain unresolved.",
        "",
        "## Diagnostics",
        "",
    ]
    for item in diagnostics:
        lines.append(
            f"- {item['arm']}: feedback total in output workbook `{item['feedback_total_from_output_excel']}`, feedback total in linkage report `{item['feedback_total_from_report']}`, tagged-message total in chat export `{item['tagged_message_total_from_report']}`, unresolved tagged messages `{item['unresolved_message_count_from_report']}`."
        )
    lines.extend(["", "## Warnings", ""])
    for message in warnings or ["No warnings detected."]:
        lines.append(f"- {message}")
    lines.extend(["", "## Mediation Results", ""])
    for result in analysis_results:
        lines.append(f"### {result['outcome_title']}")
        lines.append("")
        for sentence in build_result_interpretation(result):
            lines.append(f"- {sentence}")
        lines.append(
            f"- Binary mediator sensitivity (`>=1` feedback): indirect={format_number(result['binary_mediator_sensitivity']['indirect_effect'])}, 95% CI [{format_number(result['binary_mediator_sensitivity']['ci_low'])}, {format_number(result['binary_mediator_sensitivity']['ci_high'])}], significant={result['binary_mediator_sensitivity']['significant']}."
        )
        lines.append(
            f"- Within-arm Spearman dose-response: Human rho={format_number(result['within_arm_dose_response']['human']['spearman_rho'])} (p={format_number(result['within_arm_dose_response']['human']['spearman_p_value'])}); EPS-human rho={format_number(result['within_arm_dose_response']['eps_human']['spearman_rho'])} (p={format_number(result['within_arm_dose_response']['eps_human']['spearman_p_value'])})."
        )
        lines.append("")
    lines.extend(
        [
            "## Plot Status",
            "",
            f"- {plot_status}",
            "",
            "## Interpretation Boundary",
            "",
            "- Use this file as a sensitivity layer translating the feedback-frequency hypothesis into code.",
            "- In this workflow, the mediator is the actual feedback count reconstructed from chat history rather than a derived coverage metric.",
            "",
        ]
    )
    return "\n".join(lines)


def derive_output_paths(outdir: Path, cohort: str) -> dict[str, Path]:
    prefix = f"{cohort}_feedback_mediation"
    return {
        "summary_json": outdir / f"{prefix}_summary.json",
        "report_md": outdir / f"{prefix}_report.md",
        "results_xlsx": outdir / f"{prefix}_results.xlsx",
        "plot_png": outdir / f"{prefix}_plot.png",
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run exploratory mediation analysis using actual feedback counts reconstructed from chat history."
    )
    parser.add_argument("--cohort", choices=["weight_loss", "glycemic"], required=True, help="Clinical-trial cohort.")
    parser.add_argument("--human-file", type=Path, required=True, help="Human-arm workbook augmented with tagged-message counts.")
    parser.add_argument("--eps-file", type=Path, required=True, help="EPS-human-arm workbook augmented with tagged-message counts.")
    parser.add_argument("--human-report", type=Path, required=True, help="JSON report produced by build_checkin_dataset.py for the human arm.")
    parser.add_argument("--eps-report", type=Path, required=True, help="JSON report produced by build_checkin_dataset.py for the EPS-human arm.")
    parser.add_argument("--outdir", type=Path, required=True, help="Directory for JSON/Markdown/Excel/PNG outputs.")
    parser.add_argument("--human-count-column", default=None, help="Optional override for the human-arm feedback count column.")
    parser.add_argument("--eps-count-column", default=None, help="Optional override for the EPS-human-arm feedback count column.")
    parser.add_argument("--n-boot", type=int, default=5000, help="Number of bootstrap draws for indirect-effect intervals.")
    parser.add_argument("--seed", type=int, default=42, help="Random seed for the bootstrap.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    count_candidates_human = [args.human_count_column] if args.human_count_column else HUMAN_COUNT_COLUMNS
    count_candidates_eps = [args.eps_count_column] if args.eps_count_column else EPS_COUNT_COLUMNS
    outcome_specs = WEIGHT_OUTCOME_SPECS if args.cohort == "weight_loss" else GLYCEMIC_OUTCOME_SPECS
    output_paths = derive_output_paths(args.outdir, args.cohort)

    log("Loading participant datasets ...")
    human_rows = load_arm_dataset(
        path=args.human_file,
        cohort=args.cohort,
        group_value=HUMAN_GROUP,
        count_candidates=count_candidates_human,
        arm_label="Human",
    )
    eps_rows = load_arm_dataset(
        path=args.eps_file,
        cohort=args.cohort,
        group_value=EPS_GROUP,
        count_candidates=count_candidates_eps,
        arm_label="EPS-human",
    )
    all_rows = human_rows + eps_rows

    log("Loading checkin reports ...")
    human_report = load_json(args.human_report)
    eps_report = load_json(args.eps_report)

    log("Building diagnostics ...")
    diagnostics = build_diagnostics(human_rows, eps_rows, human_report, eps_report)
    warnings = build_quality_warnings(diagnostics)

    log("Running mediation specifications ...")
    analysis_results = [
        run_mediation_spec(
            rows=all_rows,
            cohort=args.cohort,
            spec=spec,
            n_boot=args.n_boot,
            seed=args.seed + index * 100,
        )
        for index, spec in enumerate(outcome_specs)
    ]

    log("Attempting plot generation ...")
    plot_status = maybe_build_plot(
        dataset_rows=all_rows,
        cohort=args.cohort,
        primary_result=analysis_results[0],
        plot_path=output_paths["plot_png"],
    )

    json_ready_results = []
    for result in analysis_results:
        copied = dict(result)
        copied["bootstrap_draws_summary"] = {
            "n_boot": len(result["bootstrap_draws"]),
            "mean": float(np.mean(result["bootstrap_draws"])),
            "std": float(np.std(result["bootstrap_draws"], ddof=0)),
        }
        copied["binary_bootstrap_draws_summary"] = {
            "n_boot": len(result["binary_bootstrap_draws"]),
            "mean": float(np.mean(result["binary_bootstrap_draws"])),
            "std": float(np.std(result["binary_bootstrap_draws"], ddof=0)),
        }
        copied.pop("bootstrap_draws")
        copied.pop("binary_bootstrap_draws")
        json_ready_results.append(copied)

    summary = {
        "analysis_scope": {
            "label": "exploratory mediation",
            "cohort": args.cohort,
            "description": "Actual feedback counts reconstructed from chat-export records are used as the mediator.",
            "bootstrap_iterations": args.n_boot,
            "random_seed": args.seed,
            "inputs": {
                "human_file": str(args.human_file),
                "eps_file": str(args.eps_file),
                "human_report_file": str(args.human_report),
                "eps_report_file": str(args.eps_report),
            },
        },
        "diagnostics": diagnostics,
        "warnings": warnings,
        "analysis_results": json_ready_results,
        "plot_status": plot_status,
        "outputs": {key: str(value) for key, value in output_paths.items()},
    }

    ensure_parent_dir(output_paths["summary_json"])
    output_paths["summary_json"].write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"Wrote {output_paths['summary_json'].name}")

    output_paths["report_md"].write_text(
        build_markdown_report(
            cohort=args.cohort,
            diagnostics=diagnostics,
            warnings=warnings,
            analysis_results=analysis_results,
            plot_status=plot_status,
        ),
        encoding="utf-8",
    )
    log(f"Wrote {output_paths['report_md'].name}")

    write_results_workbook(
        cohort=args.cohort,
        diagnostics=diagnostics,
        warnings=warnings,
        dataset_rows=all_rows,
        analysis_results=analysis_results,
        plot_status=plot_status,
        output_xlsx=output_paths["results_xlsx"],
    )
    log(f"Wrote {output_paths['results_xlsx'].name}")


if __name__ == "__main__":
    main()
